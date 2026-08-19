#!/usr/bin/env python3
"""
產出的 .xlsx 逐本體檢 —— 階段 5 的人工核對助手。

不是單元測試：這裡檢的是「打開這本活頁簿的人會不會被誤導」，所以查的是
分頁齊不齊、缺值是不是寫 n/a（不是 0）、關鍵指標是不是公式、分部合計公式
有沒有逐欄扣掉上層，並把幾個已知該對的數字印出來給人眼看。

  python tools/xlsx_audit.py tools/sweep_out/xlsx/*.xlsx
"""
import io
import sys
from openpyxl import load_workbook

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
else:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

EXPECT = ["說明", "損益表", "資產負債表", "現金流量表", "關鍵指標", "原始資料"]
OPTIONAL = ["估值倍數", "分部數據"]


def audit(path: str) -> None:
    wb = load_workbook(path)  # 保留公式
    name = path.replace("\\", "/").rsplit("/", 1)[-1]
    print("=" * 74)
    print(f"{name}   分頁：{wb.sheetnames}")

    for s in EXPECT:
        if s not in wb.sheetnames:
            print(f"  ✗ 缺分頁 {s}")

    zeros, nas, formulas, cells = 0, 0, 0, 0
    for s in ("損益表", "資產負債表", "現金流量表"):
        if s not in wb.sheetnames:
            continue
        ws = wb[s]
        for row in ws.iter_rows(min_row=2, min_col=3):
            for c in row:
                if c.value is None:
                    continue
                cells += 1
                if c.value == "n/a":
                    nas += 1
                elif c.value == 0:
                    zeros += 1
    print(f"  三表：{cells} 格　n/a {nas}　寫 0 的 {zeros}（0 只該出現在真的是零的科目）")

    if "關鍵指標" in wb.sheetnames:
        ws = wb["關鍵指標"]
        f = n = 0
        for row in ws.iter_rows(min_row=2, min_col=3):
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    f += 1
                elif isinstance(c.value, (int, float)):
                    n += 1
        print(f"  關鍵指標：公式 {f} 格、寫死數值 {n} 格（規格要求全公式，寫死應為 0）")

    if "估值倍數" in wb.sheetnames:
        ws = wb["估值倍數"]
        na = sum(1 for r in ws.iter_rows(min_row=25) for c in r if c.value == "n/a")
        pr = None
        for r in ws.iter_rows(min_row=25, max_col=1):
            if r[0].value == "期末股價":
                pr = r[0].row
                break
        vals = []
        if pr:
            vals = [c.value for c in ws[pr][2:] if c.value is not None]
        print(f"  估值倍數：n/a {na} 格；股價列樣本 {vals[-4:]}")
    else:
        print("  估值倍數：未產生（無股價或無股數）")

    if "分部數據" in wb.sheetnames:
        ws = wb["分部數據"]
        tot = [c for r in ws.iter_rows(max_col=1) for c in r if c.value == "　合計"]
        sample = None
        if tot:
            rr = tot[0].row
            sample = next((c.value for c in ws[rr][2:] if isinstance(c.value, str)), None)
        warn = sum(1 for r in ws.iter_rows() for c in r
                   if c.fill and c.fill.fgColor and c.fill.fgColor.rgb
                   and str(c.fill.fgColor.rgb).endswith("F7EADF"))
        print(f"  分部數據：{ws.max_row} 列、合計列 {len(tot)} 條，合計公式樣本 {sample}")
        print(f"           未校驗標色格 {warn}")
    else:
        print("  分部數據：未產生（該公司未揭露分部）")


if __name__ == "__main__":
    for p in sys.argv[1:]:
        audit(p)
