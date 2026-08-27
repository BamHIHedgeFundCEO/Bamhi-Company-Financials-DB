# -*- coding: utf-8 -*-
"""情景估值分頁：五個關卡、空輸入不變成 0、機率檢核、敏感度表、年度發行人 TTM。
python test_scenario.py
"""
import random
from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from config_loader import xbrl_map
from workbook import build_workbook

cmap = xbrl_map()


def make_payload(annual=False, quarters=16):
    random.seed(11)
    if annual:
        periods = [f"FY{2019 + i}" for i in range(8)]
    else:
        periods = [f"FY{2023 + i // 4} Q{i % 4 + 1}" for i in range(quarters)]
    line_items = []
    for c in [c for c in cmap["concepts"] if not c.get("internal")]:
        values = {}
        for i, p in enumerate(periods):
            base = 1e9 if c["unit"] == "USD" else (5e9 if c["unit"] == "shares" else 1.2)
            # 期末日：季度取 3/6/9/12 月底、年度取 12 月底
            if annual:
                end = f"{2019 + i}-12-31"
            else:
                y, q = 2023 + i // 4, i % 4 + 1
                end = f"{y}-{q * 3:02d}-{[31, 30, 30, 31][q - 1]:02d}"
            values[p] = {"value": round(base * (1 + 0.05 * i) * random.uniform(0.9, 1.1), 2),
                         "isEstimated": False, "sourceTag": c["tags"][0],
                         "accessionOrForm": "10-K", "filed": "2026-02-01", "endDate": end}
        line_items.append({**{k: c[k] for k in ("id", "zh", "en", "statement", "unit", "sign")},
                           "sourceTag": c["tags"][0], "values": values})
    valuation = {"currentPrice": 100.0, "rows": [
        {"id": "price", "values": {p: round(50 + 3 * i, 2) for i, p in enumerate(periods)}},
        {"id": "ps_vs_median", "desc": "目前 1.2 倍 歷史中位數"},
    ]}
    return {"cacheKey": "SCEN.xlsx", "financials": {
        "company": "Scenario Corp", "ticker": "SCEN", "cik": "0000000000",
        "mapVersion": cmap["version"], "periods": periods, "lookbackCount": 0,
        "periodicity": "annual" if annual else "quarterly",
        "lineItems": line_items, "derived": cmap["derived"], "valuation": valuation,
    }}


# ─────────────────────────────────────────────────────────────
data = build_workbook(make_payload())
open("scentest_out.xlsx", "wb").write(data)
wb = load_workbook(BytesIO(data))
assert "情景估值" in wb.sheetnames, wb.sheetnames
ws = wb["情景估值"]

# 1) 目標年 ＝ 最近一個會計年度（Q4）+ 3 年；不是最新一季 + 3
#    測資最後一欄是 FY2026 Q4（期末 2026-12-31）→ 目標 2029-12-31
assert ws["C25"].value.date() == date(2029, 12, 31), ws["C25"].value
assert "FY2029" in ws["A2"].value or "FY2029" in ws["A3"].value, ws["A3"].value
assert ws["C26"].value == '=IFERROR(YEARFRAC(TODAY(),$C$25),"n/a")', ws["C26"].value

# 2) 三情境欄都在，機率預設 25/50/25 且是輸入格（藍字黃底）
assert [ws.cell(row=4, column=c).value for c in (3, 4, 5)] == ["樂觀 Bull", "中性 Base", "悲觀 Bear"]
assert [ws.cell(row=5, column=c).value for c in (3, 4, 5)] == [0.25, 0.50, 0.25]
assert ws["C5"].font.color.rgb.endswith("0000FF"), ws["C5"].font.color.rgb

# 3) 五個關卡的公式串得起來，且**每一格都擋空輸入**（COUNT guard），
#    否則 Excel 把空白當 0 → EBITDA 0 → 目標價 $0.00 → 上漲空間 −100%
chain = {
    10: "$C$7*$C$9",                                  # ① 營收 × 利潤率
    13: "$C$10*$C$12",                                # ② EBITDA × 倍數
    18: "$C$13+$C$15-$C$16+N($C$17)",                 # ③ 股權橋樑
    22: "$C$18/$C$20",                                # ④ ÷ 稀釋股數
    27: "$C$22/(1+$C$24)^$C$26",                      # ⑤ 折現
}
for r, expr in chain.items():
    v = ws.cell(row=r, column=3).value
    assert v.startswith("=IF(COUNT("), (r, v)
    assert expr in v, (r, v, expr)
    assert '"待輸入"' in v, (r, v)

# 4) 輸入格真的是空的（不預填歷史值），但歷史參考欄有公式
for r in (7, 9, 12, 15, 16, 20):
    assert ws.cell(row=r, column=3).value is None, (r, ws.cell(row=r, column=3).value)
    assert ws.cell(row=r, column=4).value is None, r
for r, frag in ((7, "SUM("), (8, "^(1/3)-1"), (9, "MEDIAN("), (12, "MEDIAN("),
                (15, "IFERROR("), (16, "+"), (20, "IFERROR(")):
    v = ws.cell(row=r, column=6).value
    assert isinstance(v, str) and frag in v, (r, v)
# 認股權證那一列沒有參考值可給（10-K 附註是 HTML，數字不從 HTML 取）
assert "n/a" in ws.cell(row=17, column=6).value, ws.cell(row=17, column=6).value

# 5) 歷史參考欄真的指到別的分頁（不是自己頁內亂指）
assert "'損益表'!" in ws["F7"].value, ws["F7"].value
assert "'關鍵指標'!" in ws["F9"].value, ws["F9"].value
assert "'估值倍數'!" in ws["F12"].value, ws["F12"].value
assert "'資產負債表'!" in ws["F15"].value, ws["F15"].value

# 6) 機率合計檢核：加權兩格在 ≠100% 時不出數字
for r in (30, 31):
    v = ws.cell(row=r, column=3).value
    assert "ABS(SUM($C$5:$E$5)-1)<0.0001" in v, (r, v)
    assert "SUMPRODUCT($C$5:$E$5" in v, (r, v)
assert "TEXT(SUM($C$5:$E$5)" in ws["F5"].value, ws["F5"].value

# 7) 目前股價跨表引用估值倍數的輸入格；上漲空間有紅綠條件式
assert ws["C32"].value == "='估值倍數'!$B$3"
assert any("C33" in str(r.sqref) for r in ws.conditional_formatting)

# 8) 敏感度表：5×5，中央格用中性情境，角落用 ±2
assert ws["E41"].value == '=IF(COUNT($D$9)<1,"待輸入",$D$9+0)', ws["E41"].value
assert ws["C41"].value == '=IF(COUNT($D$9)<1,"待輸入",$D$9-0.02)', ws["C41"].value
assert ws["B44"].value == '=IF(COUNT($D$12)<1,"待輸入",$D$12+0)', ws["B44"].value
assert ws["B42"].value == '=IF(COUNT($D$12)<1,"待輸入",$D$12-2)', ws["B42"].value
mid = ws["E44"].value
assert "$D$7*E$41*$B44" in mid, mid
assert "/(1+$C$24)^$C$26" in mid, mid
assert mid.startswith("=IF(COUNT("), mid
# 邊角
assert ws["C42"].value.count("$B42") >= 1

# 9) 圖：長條、左軸含 0（openpyxl 不寫 min → Excel 長條預設自 0 起），
#    且金額格式不是「百萬美元」軸
assert len(ws._charts) == 1, len(ws._charts)
ch = ws._charts[0]
nf = ch.y_axis.number_format
assert getattr(nf, "formatCode", nf) == '$0.00', nf
assert ch.y_axis.scaling.min is None, ch.y_axis.scaling.min
assert "百萬美元" not in str(ch.title), ch.title

# 10) 說明分頁有這一頁的說明
info = wb["說明"]
labels = [info.cell(row=r, column=1).value for r in range(4, 40)]
assert "情景估值分頁" in labels, labels

# ── 年度發行人（20-F）：TTM 不能加四欄 ────────────────────────
data_a = build_workbook(make_payload(annual=True))
wba = load_workbook(BytesIO(data_a))
wsa = wba["情景估值"]
# 近 12 個月營收 ＝ 單一欄，不是 SUM
assert "SUM(" not in wsa["F7"].value, wsa["F7"].value
# 目標年 ＝ FY2026 + 3
assert wsa["C25"].value.date() == date(2029, 12, 31), wsa["C25"].value
# 3 年 CAGR：往前 3 欄（不是 12 欄）
vsa = wba["估值倍數"]
pe_row = None
for r in range(25, 40):
    if vsa.cell(row=r, column=1).value == "本益比 P/E":
        pe_row = r
assert pe_row, "找不到 P/E 列"
pe_formula = next(v for v in reversed([vsa.cell(row=pe_row, column=c).value
                                       for c in range(3, 12)]) if v)
assert "SUM(" not in pe_formula, pe_formula  # 年度發行人的 P/E 不能除以四年淨利

print("test_scenario OK")
