"""
依 chart_spec.json 程式化建圖。資料範圍由 n_quarters 動態計算，不寫死。
統一樣式（不用 openpyxl 預設）：固定色盤、淺灰虛線橫格線、圖例置下、
雙軸圖長條走主軸（金額）、折線走次軸（比率）。

⚠️ 系列一律用 Series(values_ref, title=中文名) 建立，不用 add_data：
add_data 預設把「每欄」當一條系列（要 from_rows），之前因此整張圖亂掉。
"""
import re

from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.utils import get_column_letter

from config_loader import theme
from formulas import FIRST_DATA_COL

# 圖表資料起始欄：預設 C；有 lookback 隱藏欄時，workbook 設成第一顯示欄，
# 讓圖表只畫顯示範圍（公式仍引用含 lookback 的完整欄位）。
_DATA_START = FIRST_DATA_COL


def set_data_start(col: int):
    global _DATA_START
    _DATA_START = col


def _hex(c: str) -> str:
    return c.lstrip("#").upper()


def _dash_line(color: str):
    gp = GraphicalProperties()
    gp.line = LineProperties(solidFill=_hex(color), prstDash="dash", w=9525)
    return gp


def _style_chart(chart, th: dict):
    ch = th["chart"]
    if chart.legend is not None:
        chart.legend.position = ch.get("legend_position", "b")
        chart.legend.overlay = False
    chart.y_axis.majorGridlines = ChartLines()
    chart.y_axis.majorGridlines.graphicalProperties = _dash_line(ch["gridline_color"])
    chart.x_axis.majorGridlines = None
    # 座標軸刻度數字必須可見（openpyxl 預設 delete/tickLblPos 會讓 Excel 隱藏刻度）
    for ax in (chart.x_axis, chart.y_axis):
        ax.delete = False
        ax.tickLblPos = "nextTo"
        ax.majorTickMark = "out"
    chart.x_axis.tickLblPos = "low"  # 有負值時 X 軸標籤仍留在下方
    # 移除繪圖區/整體粗框
    chart.graphical_properties = GraphicalProperties()
    chart.graphical_properties.line = LineProperties(noFill=True)


def _values_ref(ws, row: int, n_quarters: int) -> Reference:
    """只含數據欄（C 起），不含科目名欄。"""
    return Reference(ws, min_col=_DATA_START, max_col=_DATA_START - 1 + n_quarters,
                     min_row=row, max_row=row)


def _cats_ref(ws, n_quarters: int) -> Reference:
    return Reference(ws, min_col=_DATA_START, max_col=_DATA_START - 1 + n_quarters,
                     min_row=1, max_row=1)


def _make_series(ws, row: int, n_quarters: int, color: str, is_line: bool) -> Series:
    s = Series(_values_ref(ws, row, n_quarters), title=str(ws.cell(row=row, column=1).value or ""))
    c = _hex(color)
    gp = GraphicalProperties()
    if is_line:
        gp.line = LineProperties(solidFill=c, w=22225)  # 1.75pt
        gp.noFill = True
        s.smooth = False
    else:
        gp.solidFill = c
        gp.line = LineProperties(noFill=True)
        s.invertIfNegative = False  # 負值長條沿用同色，不要翻成白色
    s.graphicalProperties = gp
    return s


def _period_labels(cats_ws, n_quarters: int) -> list[str]:
    return [str(cats_ws.cell(row=1, column=_DATA_START + i).value or "") for i in range(n_quarters)]


def _color_bars_by_year(series: Series, labels: list[str], palette: list[str]):
    """單一科目長條圖：同一會計年度的季用同一色，不同年換色 → 一眼看出年度分界。"""
    fy_order: list[str] = []
    for lb in labels:
        m = re.match(r"FY(\d{4})", lb)
        fy = m.group(1) if m else lb
        if fy not in fy_order:
            fy_order.append(fy)
    color_of = {fy: _hex(palette[i % len(palette)]) for i, fy in enumerate(fy_order)}
    series.invertIfNegative = False  # 負值長條也要上色（否則翻白）
    pts = []
    for i, lb in enumerate(labels):
        m = re.match(r"FY(\d{4})", lb)
        fy = m.group(1) if m else lb
        gp = GraphicalProperties(solidFill=color_of[fy])
        gp.line = LineProperties(noFill=True)
        dp = DataPoint(idx=i, spPr=gp)
        dp.invertIfNegative = False
        pts.append(dp)
    series.data_points = pts


def build_chart(spec: dict, cats_ws, locate, n_quarters: int):
    """spec: chart_spec.json 單一圖表物件。locate(id) → (ws, row) 或 None（全活頁簿定位）。"""
    th = theme()
    palette = th["palette"]["chart_series"]
    ci = 0

    def series_for(ids, is_line):
        nonlocal ci
        out = []
        for i in ids:
            loc = locate(i)
            if loc:
                out.append(_make_series(loc[0], loc[1], n_quarters, palette[ci % len(palette)], is_line))
                ci += 1
        return out

    kind = spec["type"]
    cats = _cats_ref(cats_ws, n_quarters)

    if kind in ("bar", "stacked_bar"):
        chart = BarChart()
        chart.type = "col"
        chart.gapWidth = 60
        if kind == "stacked_bar":
            chart.grouping = "stacked"
            chart.overlap = 100
        chart.series = series_for(spec.get("series", []), is_line=False)
        chart.set_categories(cats)
        # 單一科目長條圖 → 依會計年度上色（成本費用結構這類多科目堆疊圖維持各科目一色）
        if len(chart.series) == 1:
            _color_bars_by_year(chart.series[0], _period_labels(cats_ws, n_quarters), palette)
        is_money = True
    elif kind == "line":
        chart = LineChart()
        chart.series = series_for(spec.get("series", []), is_line=True)
        chart.set_categories(cats)
        is_money = False
    elif kind == "bar+line":
        chart = BarChart()
        chart.type = "col"
        chart.gapWidth = 60
        chart.series = series_for(spec.get("bars", []), is_line=False)
        chart.set_categories(cats)
        if len(chart.series) == 1:
            _color_bars_by_year(chart.series[0], _period_labels(cats_ws, n_quarters), palette)
        line = LineChart()
        line.series = series_for(spec.get("line", []), is_line=True)
        line.set_categories(cats)
        if spec.get("secondary_axis"):
            line.y_axis.axId = 200
            line.y_axis.crosses = "max"
            line.y_axis.majorGridlines = None
            line.y_axis.delete = False
            line.y_axis.tickLblPos = "nextTo"
        chart += line
        is_money = True
    else:
        return None

    if not chart.series:
        return None

    # Y 軸單位：spec 的 y_format 優先（percent/days/multiple/money）；否則沿用 is_money 判斷
    MONEY = th["number_formats"]["usd_millions_axis"]
    yf = spec.get("y_format")
    fmt_map = {
        "percent": "0%",
        "days": '0"天"',
        "multiple": '0.0"x"',
        "money": MONEY,
    }
    if yf in fmt_map:
        y_fmt = fmt_map[yf]
    elif is_money:
        y_fmt = MONEY
    else:
        y_fmt = None
    is_money_axis = y_fmt == MONEY

    chart.title = spec["title"] + ("（百萬美元）" if is_money_axis else "")
    chart.width = th["chart"]["width_cols"] * 2.0    # 放大：約 24cm
    chart.height = th["chart"]["height_rows"] * 0.7  # 約 12.6cm
    if y_fmt:
        chart.y_axis.number_format = y_fmt
        chart.y_axis.numFmt = y_fmt
    # 單一系列的折線/長條 → 圖例無意義，移除
    if len(chart.series) <= 1:
        chart.legend = None
    _style_chart(chart, th)
    return chart


def place_charts(ws, specs: list[dict], locate, n_quarters: int, anchor_row: int,
                 index: list | None = None) -> int:
    """資料區下方直向排列，每張圖獨立一段，留足間距不重疊。回傳下一個可用列。
    index（若給）收集 (工作表名, 圖表標題, 錨定列) 供「說明」分頁做圖表目錄超連結。"""
    th = theme()
    # 圖高（cm）→ 佔用列數（放大後約 11cm；預設列高約 0.53cm），再加 5 列間隔
    step = int(th["chart"]["height_rows"] * 0.7 / 0.53) + 5
    r = anchor_row
    for spec in specs:
        chart = build_chart(spec, ws, locate, n_quarters)
        if chart is None:
            continue
        ws.add_chart(chart, f"{get_column_letter(_DATA_START)}{r}")
        if index is not None:
            index.append((ws.title, spec["title"], r))
        r += step
    return r
