#!/usr/bin/env python3
"""
全市場對照表缺口掃描 —— coverage.py 的批次版。

和 coverage.py 的分工：
  coverage.py  單檔深挖，印給人看
  sweep.py     上千檔掃描，產出可 diff 的 JSONL + 跨公司彙總的候選標籤

**只回答「config 缺標籤」這一類 n/a。** 期間對不上、Q4 推算錯這類要打真的 API，
不在這裡模擬 —— 自己在 Python 重寫一份管線，量到的是那份重寫，不是網站真的吐的。

用法：
  python tools/sweep.py --universe 羅素1000.xlsx        # 全掃（首次約 14 分鐘 / 3.4GB）
  python tools/sweep.py --universe ... --limit 50       # 先試 50 檔
  python tools/sweep.py --report                        # 只讀既有結果重新彙總（秒級）

companyfacts 快取在 ~/.bamhi-facts-cache（gzip，約 400MB）。已申報財報不可變，
快取永久有效；砍掉只是要重抓。改完 config 重跑會直接吃快取，不再打 SEC。
"""
import argparse
import gzip
import io
import json
import os
import re
import sys
import time
import urllib.error
import urllib.request
from collections import Counter, defaultdict

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
else:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(ROOT, "tools"))
from coverage import NAMESPACES, suggest_tags, tag_has_data, unit_prefs  # noqa: E402

UA = os.environ.get("SEC_USER_AGENT", "BamHI frank940702@gmail.com")
MAP_PATH = os.path.join(ROOT, "config", "xbrl_zh_map.json")
CACHE = os.environ.get("BAMHI_FACTS_CACHE",
                       os.path.join(os.path.expanduser("~"), ".bamhi-facts-cache"))
OUT = os.path.join(ROOT, "tools", "sweep_out")


def get_json(url: str) -> dict:
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req) as r:
        return json.load(r)


def load_universe(path: str) -> list[dict]:
    """彭博格式的成分股表：`AAPL UW Equity` / 產業 / 細產業。"""
    from openpyxl import load_workbook

    ws = load_workbook(path, read_only=True).worksheets[0]
    out = []
    for r in ws.iter_rows(values_only=True):
        if not r or not r[0]:
            continue
        m = re.match(r"^(\S+)\s+\S\S\s+Equity$", str(r[0]).strip())
        if not m:
            print(f"格式看不懂，略過：{r[0]}")
            continue
        out.append({"ticker": m.group(1).replace("/", "."),
                    "sector": r[1] if len(r) > 1 else "",
                    "industry": r[2] if len(r) > 2 else ""})
    return out


def build_cik_lut() -> dict[str, str]:
    """SEC 兩份名冊合併。兩份都會漏（實測缺 AEP），之後靠 submissions 補。"""
    lut = {}
    for v in get_json("https://www.sec.gov/files/company_tickers.json").values():
        lut[v["ticker"].upper()] = str(v["cik_str"]).zfill(10)
    ex = get_json("https://www.sec.gov/files/company_tickers_exchange.json")
    ti, ci = ex["fields"].index("ticker"), ex["fields"].index("cik")
    for row in ex["data"]:
        if row[ti]:
            lut.setdefault(str(row[ti]).upper(), str(row[ci]).zfill(10))
    return lut


def resolve(ticker: str, lut: dict[str, str]) -> str | None:
    # SEC 名冊用連字號（BRK-B），成分股表用點（BRK.B / BRK/B）
    for k in (ticker, ticker.replace(".", "-"), ticker.replace("-", ".")):
        cik = lut.get(k.upper())
        if cik:
            return cik
    return None


def facts_of(cik: str) -> dict | None:
    """companyfacts，優先讀本機快取。回傳 facts 節點；抓不到回 None。"""
    os.makedirs(CACHE, exist_ok=True)
    path = os.path.join(CACHE, f"CIK{cik}.json.gz")
    if os.path.exists(path):
        with gzip.open(path, "rt", encoding="utf-8") as f:
            return json.load(f).get("facts")
    url = f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": UA})
        with urllib.request.urlopen(req) as r:
            raw = r.read()
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return None
        raise
    with gzip.open(path, "wt", encoding="utf-8") as f:
        f.write(raw.decode("utf-8"))
    time.sleep(0.12)  # SEC 限速；讀快取時不 sleep
    return json.loads(raw).get("facts")


def annual_ends(facts: dict, tag: str, prefs: list[str]) -> int:
    """該標籤有幾個「年度」期間（340–400 天）—— 只有一兩期一樣會滿版 n/a。"""
    from datetime import date

    ends = set()
    for ns in NAMESPACES:
        node = facts.get(ns, {}).get(tag)
        if not node:
            continue
        for u in prefs:
            for p in node.get("units", {}).get(u, []):
                if not p.get("start"):  # 存量科目（資產負債表）無 start，用期末日算
                    ends.add(p["end"][:4])
                    continue
                n = (date.fromisoformat(p["end"]) - date.fromisoformat(p["start"])).days
                if 340 <= n <= 400:
                    ends.add(p["end"][:4])
    return len(ends)


def scan_one(rec: dict, cik: str, concepts: list[dict]) -> dict | None:
    facts = facts_of(cik)
    if facts is None:
        return None
    is_ifrs = "us-gaap" not in facts or len(facts.get("us-gaap", {})) < 20

    hits: dict[str, dict] = {}
    for c in concepts:
        tags = c.get("tags_ifrs", []) if is_ifrs else c["tags"]
        prefs = unit_prefs(c["unit"])
        for tag in tags:
            ok, end, val = tag_has_data(facts, tag, prefs)
            if ok:
                hits[c["id"]] = {"tag": tag, "latest": end,
                                 "years": annual_ends(facts, tag, prefs)}
                break

    # 管線的兩層 fallback 也要算進來，否則會把「其實有值」誤報成缺
    has_bs = "total_assets" in hits
    derived, zeroed = [], []
    # 推算可以串接：管線照 concepts 順序跑，opex_total 的 gross_profit 本身
    # 常常也是推算來的（營收−成本）。只認直接標籤會誤報 200 家以上。
    # 用不動點迭代而不是照順序掃一次，config 換順序也不會壞。
    resolved_now = set(hits)
    for _ in range(len(concepts)):
        grew = False
        for c in concepts:
            if c["id"] in resolved_now or not c.get("derive"):
                continue
            deps = re.findall(r"[a-z_][a-z0-9_]+", c["derive"])
            if all(d in resolved_now for d in deps):
                derived.append(c["id"])
                resolved_now.add(c["id"])
                grew = True
        if not grew:
            break
    for c in concepts:
        if c["id"] in hits or c["id"] in resolved_now:
            continue
        if c.get("zero_if_absent") and has_bs:
            zeroed.append(c["id"])
            continue
        # 第三層退路：期末股數缺就用加權平均近似（financials.ts 的
        # shares_outstanding 補齊段）。這是寫在程式裡的行為，config 看不出來，
        # 不模擬的話會把 75 家多股別公司誤報成缺。
        if c["id"] == "shares_outstanding" and (
                "shares_basic" in hits or "shares_diluted" in hits):
            derived.append(c["id"])

    resolved = set(hits) | set(derived) | set(zeroed)
    missing = [c["id"] for c in concepts if c["id"] not in resolved]

    # 缺的科目就地找候選標籤 —— 跨公司彙總後才看得出哪個值得加進 config
    #
    # 兩種強度分開存，因為可信度差很多：
    #   kin  同族標籤：公司的標籤 = config 標籤**加後綴**。ADBE 的
    #        ResearchAndDevelopmentExpenseSoftwareExcludingAcquiredInProcessCost
    #        就是這樣抓到的（= ResearchAndDevelopmentExpense + 後綴），可信。
    #
    #        只認這一個方向。反過來（公司標籤比較短）全是雜訊：Liabilities 之於
    #        LiabilitiesCurrent —— 總負債不是流動負債，採用就是拿錯數字。加前綴
    #        的也不行：OtherRestrictedAssetsCurrent 不是流動資產合計。後綴通常是
    #        同一科目的細分修飾，前綴則是換成另一個科目。
    #   cands 關鍵詞比對（suggest_tags）：只能當線索，會把 InterestExpense 當研發。
    cands: dict[str, list] = {}
    kin: dict[str, list] = {}
    lib = {t for ns in NAMESPACES for t in facts.get(ns, {})}
    for c in concepts:
        if c["id"] not in missing:
            continue
        known = set(c.get("tags_ifrs", []) if is_ifrs else c["tags"])
        cands[c["id"]] = [t for t, _e, _v in suggest_tags(facts, c, known)]
        rel = [t for t in lib if t not in known
               and any(t.startswith(k) for k in known)]
        # 只留真的有數字的，否則會建議一堆空標籤
        kin[c["id"]] = sorted(
            t for t in rel if tag_has_data(facts, t, unit_prefs(c["unit"]))[0]
        )

    return {"ticker": rec["ticker"], "cik": cik, "sector": rec["sector"],
            "industry": rec["industry"], "ifrs": is_ifrs, "n_tags_total":
            sum(len(facts.get(ns, {})) for ns in NAMESPACES),
            "hits": hits, "derived": derived, "zeroed": zeroed,
            "missing": missing, "candidates": cands, "kin": kin}


# 同業內缺這麼高比例 → 判定為結構性不適用（銀行沒存貨、航空沒研發、REIT 沒毛利）。
# 用資料自己判，不硬編碼「金融業沒有存貨」這種規則 —— 硬編碼會把真的壞掉的一起蓋掉。
#
# 產業（Financial）與細產業（Banks / REITS / Airlines）兩層都看：只看產業會漏掉
# 「航空公司沒有研發」（Consumer, Cyclical 整體只缺 76%，過不了門檻，310 家全被
# 丟進待查清單）；只看細產業則有些細產業家數太少，比例沒有意義。
STRUCTURAL = 0.80
MIN_PEERS = 8   # 產業層
MIN_PEERS_IND = 5  # 細產業層


def report(rows: list[dict], concepts: list[dict]) -> None:
    zh = {c["id"]: c["zh"] for c in concepts}
    n = len(rows)
    groups = {"sector": (Counter(r["sector"] for r in rows), MIN_PEERS),
              "industry": (Counter(r["industry"] for r in rows), MIN_PEERS_IND)}
    miss_by: dict[str, dict[str, Counter]] = {k: defaultdict(Counter) for k in groups}
    miss_total = Counter()
    for r in rows:
        for cid in r["missing"]:
            miss_total[cid] += 1
            for k in groups:
                miss_by[k][cid][r[k]] += 1

    def rate(kind: str, cid: str, g: str) -> float:
        tot = groups[kind][0][g]
        return miss_by[kind][cid][g] / tot if tot else 0.0

    structural: dict[str, dict[str, list[str]]] = {k: defaultdict(list) for k in groups}
    for cid in miss_total:
        for kind, (counts, floor) in groups.items():
            for g, tot in counts.items():
                if tot >= floor and rate(kind, cid, g) >= STRUCTURAL:
                    structural[kind][g].append(cid)

    def is_structural(cid: str, r: dict) -> bool:
        return any(cid in structural[k][r[k]] for k in groups)

    print(f"\n{'='*78}\n掃描 {n} 家 · 科目 {len(concepts)} 個")
    print(f"完全沒缺口的公司 {sum(1 for r in rows if not r['missing'])} 家")

    # 段 1：同族標籤 —— 這才是工作清單。公司用的標籤是 config 標籤的變體
    # （ADBE 的 ...ExpenseSoftwareExcluding... 之於 ResearchAndDevelopmentExpense），
    # 幾乎必然是同一個科目，加進 config 就能修。
    kin_votes: dict[str, Counter] = defaultdict(Counter)
    kin_firms: dict[tuple[str, str], list[str]] = defaultdict(list)
    for r in rows:
        for cid, tags in r.get("kin", {}).items():
            for t in tags:
                kin_votes[cid][t] += 1
                kin_firms[(cid, t)].append(r["ticker"])

    print(f"\n{'='*78}\n段 1 · 同族標籤缺口（config 漏收的標籤變體 → 可直接修）\n")
    found = False
    for cid, votes in sorted(kin_votes.items(), key=lambda x: -sum(x[1].values())):
        for tag, v in votes.most_common(4):
            found = True
            firms = kin_firms[(cid, tag)]
            print(f"  {cid:22s} {zh.get(cid,''):12s} + {tag}")
            print(f"      {v:4d} 家   {' '.join(firms[:12])}" + (" …" if v > 12 else ""))
    if not found:
        print("  （無）")

    print(f"\n{'='*78}\n段 2 · 結構性不適用（同業 ≥{STRUCTURAL:.0%} 都沒有 → n/a 正確，不要動）\n")
    for kind in groups:
        for g, cids in sorted(structural[kind].items(),
                              key=lambda x: -groups[kind][0][x[0]]):
            print(f"  [{kind}] {g} （{groups[kind][0][g]} 家）："
                  + "、".join(f"{zh.get(c, c)}{rate(kind, c, g):.0%}"
                              for c in sorted(cids, key=lambda c: -rate(kind, c, g))))
    if not any(structural[k] for k in groups):
        print("  （無）")

    # 段 3：同業多數都有、這幾家卻缺，而且不是同族標籤能解釋的
    suspects: list[tuple[int, str, list[dict]]] = []
    for cid in miss_total:
        bad = [r for r in rows if cid in r["missing"] and not is_structural(cid, r)
               and not r.get("kin", {}).get(cid)]
        if bad:
            suspects.append((len(bad), cid, bad))
    suspects.sort(reverse=True, key=lambda x: x[0])

    print(f"\n{'='*78}\n段 3 · 其餘可疑缺口（同業多數都有，且沒有同族標籤可解釋）\n")
    for cnt, cid, bad in suspects[:15]:
        peers = ", ".join(f"{g}{rate('industry', cid, g):.0%}" for g, _ in
                          Counter(r["industry"] for r in bad).most_common(3))
        print(f"  {cid:22s} {zh.get(cid,''):12s} {cnt:4d} 家   細產業缺口率 {peers}")
        print(f"      公司：{' '.join(r['ticker'] for r in bad[:16])}"
              + (f" …共 {cnt}" if cnt > 16 else ""))
        votes = Counter()
        for r in bad:
            for t in r["candidates"].get(cid, [])[:3]:
                votes[t] += 1
        for tag, v in votes.most_common(2):
            print(f"      關鍵詞候選（雜訊大）{tag:<44s} {v:4d}/{cnt}")
        if not votes:
            print("      （這些公司標籤庫也找不到相近標籤 → 很可能真的沒揭露）")

    # 有標籤但年度期數太少 → 網站舊欄位一樣是 n/a，成因不同，修法也不同
    print(f"\n{'='*78}\n段 C · 有標籤但年度期數不足（新上市或改標籤 → 舊欄位 n/a）\n")
    thin = Counter()
    for r in rows:
        for cid, h in r["hits"].items():
            if h.get("years", 0) < 3:
                thin[cid] += 1
    for cid, cnt in thin.most_common(12):
        print(f"  {cid:24s} {zh.get(cid,''):14s} {cnt:4d} 家只有 <3 個年度")

    print(f"\n{'='*78}")
    print("讀法：段 1 是工作清單，加進 config 即可修 —— 但仍要開該公司 companyfacts")
    print("      確認語意（同族不等於同義，Gross 與 Net 差很多）。")
    print("      段 2 不要動：改了就是把公司沒揭露的東西編出來。")
    print("      段 3 多半是真的沒揭露；關鍵詞候選雜訊很大（會把 InterestExpense")
    print("      當研發費用），只能當線索，不可照抄。")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--universe", help="成分股 xlsx（彭博格式）")
    ap.add_argument("--limit", type=int, help="只掃前 N 檔")
    ap.add_argument("--report", action="store_true", help="不掃描，只重讀結果彙總")
    args = ap.parse_args()

    xbrl_map = json.load(open(MAP_PATH, encoding="utf-8"))
    concepts = xbrl_map["concepts"]
    os.makedirs(OUT, exist_ok=True)
    out_path = os.path.join(OUT, "coverage.jsonl")

    if args.report:
        rows = [json.loads(ln) for ln in open(out_path, encoding="utf-8")]
        report(rows, concepts)
        return

    if not args.universe:
        ap.error("要 --universe 檔案，或用 --report 讀既有結果")

    uni = load_universe(args.universe)
    if args.limit:
        uni = uni[: args.limit]
    lut = build_cik_lut()

    rows, unresolved, nofacts = [], [], []
    t0 = time.time()
    for i, rec in enumerate(uni, 1):
        cik = resolve(rec["ticker"], lut)
        if not cik:
            unresolved.append(rec["ticker"])
            continue
        try:
            row = scan_one(rec, cik, concepts)
        except Exception as e:  # 單檔壞掉不能中斷整輪掃描
            print(f"  {rec['ticker']}: {type(e).__name__} {e}")
            continue
        if row is None:
            nofacts.append(rec["ticker"])
            continue
        rows.append(row)
        if i % 50 == 0:
            print(f"  {i}/{len(uni)} 已掃 {len(rows)} 家 · {time.time()-t0:.0f}s")

    with open(out_path, "w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    print(f"\n掃完 {len(rows)} 家，{time.time()-t0:.0f}s → {out_path}")
    if unresolved:
        print(f"解析不到 CIK（{len(unresolved)}）：{' '.join(unresolved)}")
    if nofacts:
        print(f"有 CIK 但無 companyfacts（{len(nofacts)}）：{' '.join(nofacts)}")
    report(rows, concepts)


if __name__ == "__main__":
    main()
