"""煙霧測試：合成 6 季資料 → 生成活頁簿 → 重新讀回驗證硬規則。
python test_workbook.py
"""
import random

from openpyxl import load_workbook
from io import BytesIO

from config_loader import xbrl_map
from workbook import build_workbook

random.seed(7)
periods = [f"FY2025 Q{q}" for q in (1, 2, 3, 4)] + [f"FY2026 Q{q}" for q in (1, 2)]

cmap = xbrl_map()
line_items = []
# internal 科目不會出現在 API 回應裡（financials.ts 最後濾掉），測資也不能有
for c in [c for c in cmap["concepts"] if not c.get("internal")]:
    values = {}
    for i, p in enumerate(periods):
        if c["id"] == "inventory" and i < 2:
            continue  # 故意留缺值 → 應顯示 n/a
        base = 1e9 if c["unit"] == "USD" else (5e9 if c["unit"] == "shares" else 1.2)
        values[p] = {
            "value": round(base * (1 + 0.1 * i) * random.uniform(0.8, 1.2), 2),
            "isEstimated": p.endswith("Q4"),
            "sourceTag": c["tags"][0],
            "accessionOrForm": "10-Q",
            "filed": "2026-05-01",
            "endDate": "2026-04-30",
        }
    line_items.append({**{k: c[k] for k in ("id", "zh", "en", "statement", "unit", "sign")},
                       "sourceTag": c["tags"][0], "values": values})

payload = {
    "cacheKey": "TEST_2025Q1_2026Q2_0.1.xlsx",
    "financials": {
        "company": "Test Corp", "ticker": "TEST", "cik": "0000000000",
        "mapVersion": cmap["version"], "periods": periods,
        "lineItems": line_items, "derived": cmap["derived"],
    },
}

data = build_workbook(payload)
open("test_out.xlsx", "wb").write(data)
wb = load_workbook(BytesIO(data))

assert wb.sheetnames == ["說明", "損益表", "資產負債表", "現金流量表", "關鍵指標", "原始資料"], wb.sheetnames

is_ws = wb["損益表"]
assert is_ws.freeze_panes == "C2"
assert is_ws["A1"].value == "科目" and is_ws["C1"].value == "FY2025 Q1"

bs = wb["資產負債表"]
inv_row = next(r for r in range(2, 40) if bs.cell(r, 1).value == "存貨")
assert bs.cell(inv_row, 3).value == "n/a", "缺值必須是 n/a"

m = wb["關鍵指標"]
gm_row = next(r for r in range(2, 60) if m.cell(r, 1).value == "毛利率")
f = m.cell(gm_row, 3).value
assert isinstance(f, str) and f.startswith("=IFERROR("), f
assert "損益表" in f
assert m.cell(gm_row, 1).comment is not None, "指標名稱要有 hover 註解"

yoy_row = next(r for r in range(2, 60) if m.cell(r, 1).value == "營收年增率")
# 比較基期落在所選區間之外 → 「—」（不適用），不是 n/a（該有卻查不到）。
# 這條在 workbook.py 判「不適用」時就分流了，測試曾停留在舊語意。
assert m.cell(yoy_row, 3).value == "—", "比較基期不在區間內時 YoY 應為「—」"
assert str(m.cell(yoy_row, 3 + 4).value).startswith("=IFERROR("), m.cell(yoy_row, 7).value

roe_row = next(r for r in range(2, 60) if m.cell(r, 1).value == "股東權益報酬率")
assert "資產負債表" in str(m.cell(roe_row, 4).value)

# 圖表存在
assert len(is_ws._charts) >= 3, f"損益表應有 3 張圖，得 {len(is_ws._charts)}"
assert len(m._charts) >= 4

print("OK — 分頁、凍結、n/a、公式、註解、圖表全過")
print("公式範例（毛利率）:", f)
print("公式範例（YoY）:", m.cell(yoy_row, 7).value)
