"""
指標公式翻譯器：xbrl_zh_map.json 的 formula 字串 → Excel 公式。
關鍵指標分頁必須寫公式不能寫算好的數值（使用者改假設整張表要能重算）。

支援語法：
  identifier            → 該 concept / 前面已定義 metric 的同欄儲存格
  identifier[t-4]       → 往前推 4 欄（不足時整條公式回 None → n/a）
  avg(identifier)       → (本欄 + 前一欄)/2；無前一欄時退回本欄
  + - * / ( ) 數字
整條公式一律包 IFERROR(..., "n/a")（除零與引用 n/a 文字格皆回 n/a）。
"""
import re
from openpyxl.utils import get_column_letter

_TOKEN = re.compile(r"([A-Za-z_][A-Za-z0-9_]*)(\[t(?:-(\d+))?\])?|(\d+\.?\d*)|([()+\-*/])|(\s+)")

FIRST_DATA_COL = 3  # C 欄起為季度


class RefResolver:
    """id → (工作表名, 列號)。concepts 指向三大報表分頁；先前定義的 metric 指向關鍵指標分頁。"""

    def __init__(self):
        self.refs: dict[str, tuple[str, int]] = {}

    def add(self, item_id: str, sheet: str, row: int):
        self.refs[item_id] = (sheet, row)

    def cell(self, item_id: str, col: int) -> str | None:
        if item_id not in self.refs or col < FIRST_DATA_COL:
            return None
        sheet, row = self.refs[item_id]
        return f"'{sheet}'!{get_column_letter(col)}{row}"


def translate(formula: str, resolver: RefResolver, col: int, annual: bool = False) -> str | None:
    """回傳不含開頭 = 的公式；引用解析失敗（缺科目或期數不足）回 None。

    annual=True（IFRS 年度模式）時：
      - [t-4]（去年同季）→ 前 1 欄；[t-1]（上一季）無意義 → None
      - 「× 4 年化」係數改 × 1；週轉天數 91.25 → 365
    """
    if annual:
        if "[t-1]" in formula:
            return None  # QoQ 類指標年度資料無意義
        formula = re.sub(r"\*\s*4\b", "* 1", formula).replace("91.25", "365")
    out: list[str] = []
    pos = 0
    while pos < len(formula):
        m = _TOKEN.match(formula, pos)
        if not m:
            return None
        ident, bracket, lag, num, op, ws = m.groups()
        pos = m.end()
        if ws:
            continue
        if num:
            out.append(num)
        elif op:
            out.append(op)
        elif ident == "avg":
            # avg(x) → (x本欄 + x前一欄)/2
            m2 = re.match(r"\(\s*([A-Za-z_][A-Za-z0-9_]*)\s*\)", formula[pos:])
            if not m2:
                return None
            pos += m2.end()
            cur = resolver.cell(m2.group(1), col)
            prev = resolver.cell(m2.group(1), col - 1)
            if cur is None:
                return None
            out.append(f"(({cur}+{prev})/2)" if prev else cur)
        elif ident:
            lag_n = int(lag) if lag else 0
            if annual and lag_n:
                lag_n = lag_n // 4  # 年度模式：[t-4] = 前一年 = 前 1 欄
            ref = resolver.cell(ident, col - lag_n)
            if ref is None:
                return None
            out.append(ref)
        else:
            return None
    expr = "".join(out)
    # 一律包 IFERROR：除零回 n/a，且引用到 "n/a" 文字格的加減式（如 FCF = CFO − CapEx）
    # 會回 #VALUE!，也要吃掉
    return f'IFERROR({expr},"n/a")'
