# -*- coding: utf-8 -*-
"""估值分頁橫向版面 + 目標價修正 + ignoredErrors 驗證。
python test_valuation_layout.py
"""
import random, zipfile, re
from io import BytesIO
from openpyxl import load_workbook

from config_loader import xbrl_map
from workbook import build_workbook

random.seed(7)
# 8 期、lookback 4 → 顯示 4 季，data_start=G（模擬 NVDA C-F 隱藏）
periods = [f"FY2025 Q{q}" for q in (1, 2, 3, 4)] + [f"FY2026 Q{q}" for q in (1, 2, 3, 4)]
cmap = xbrl_map()

line_items = []
for c in cmap["concepts"]:
    values = {}
    for i, p in enumerate(periods):
        base = 1e9 if c["unit"] == "USD" else (5e9 if c["unit"] == "shares" else 1.2)
        values[p] = {"value": round(base * (1 + 0.1 * i) * random.uniform(0.8, 1.2), 2),
                     "isEstimated": p.endswith("Q4"), "sourceTag": c["tags"][0],
                     "accessionOrForm": "10-Q", "filed": "2026-05-01", "endDate": "2026-04-30"}
    line_items.append({**{k: c[k] for k in ("id", "zh", "en", "statement", "unit", "sign")},
                       "sourceTag": c["tags"][0], "values": values})

valuation = {
    "currentPrice": 225.16,
    "rows": [
        {"id": "price", "values": {p: round(10 + 5 * i, 2) for i, p in enumerate(periods)}},
        {"id": "ps_vs_median", "desc": "目前 1.2 倍 歷史中位數"},
    ],
}

payload = {"cacheKey": "VALTEST.xlsx", "financials": {
    "company": "Test Corp", "ticker": "TEST", "cik": "0000000000",
    "mapVersion": cmap["version"], "periods": periods, "lookbackCount": 4,
    "lineItems": line_items, "derived": cmap["derived"], "valuation": valuation,
}}

data = build_workbook(payload)
open("valtest_out.xlsx", "wb").write(data)

# --- 讀回驗證 ---
wb = load_workbook(BytesIO(data))
assert "估值倍數" in wb.sheetnames, wb.sheetnames
vws = wb["估值倍數"]

# 1) 版面：三區塊橫向並排（G=data_start）
assert vws["A2"].value == "■ 假設區　Assumptions", vws["A2"].value
assert vws["G2"].value == "■ 前瞻估值　Forward Valuation", vws["G2"].value
assert vws["I2"].value == "■ 反推目標價　Reverse: Target Price", vws["I2"].value
assert vws["A3"].value == "目前股價" and vws["B3"].value == 225.16
assert vws["G3"].value == "市值" and str(vws["H3"].value).startswith("=IFERROR($B$3*$B$4")
assert vws["I5"].value == "目標價（P/S × TTM 營收）"

# 2) 目標價 bug 修正：除以 $B$4，且公式內無殘餘 "=" 造成的 "/="
tps = vws["J5"].value
assert tps is not None, "目標價 P/S 公式應存在（有 revenue）"
assert "/$B$4" in tps and "/=" not in tps, tps

# 3) 前瞻/上漲空間公式指向正確
assert vws["H4"].value == '=IFERROR($B$3/$B$5,"n/a")', vws["H4"].value
assert vws["J6"].value == '=IFERROR($J$3/$B$3-1,"n/a")', vws["J6"].value

# 4) 第 1 列（data_start 起）必須淨空 → 不污染圖表 category
for col in range(7, 11):
    assert vws.cell(row=1, column=col).value is None, f"row1 col{col} 應淨空"

# 5) 凍結窗格
assert vws.freeze_panes == "C9", vws.freeze_panes

# 6) 歷史表仍在（第 24 列標頭）
assert vws["A24"].value == "歷史逐季倍數"
assert vws.cell(row=25, column=1).value == "期末股價"

# 7) lookback 欄 C-F 隱藏
assert all(vws.column_dimensions[c].hidden for c in "CDEF"), "C-F 應隱藏"

# 8) 估值圖存在（5 張）
assert len(vws._charts) == 5, f"估值分頁應有 5 張圖，得 {len(vws._charts)}"

# 9) ignoredErrors 已注入該分頁 XML
z = zipfile.ZipFile(BytesIO(data))
sheet_no = wb.sheetnames.index("估值倍數") + 1
xml = z.read(f"xl/worksheets/sheet{sheet_no}.xml").decode("utf-8")
assert "<ignoredErrors>" in xml and "<drawing " in xml
assert xml.index("<ignoredErrors>") < xml.index("<drawing "), "ignoredErrors 必須在 <drawing> 之前"

# 10) 估值圖 category 仍讀第 1 列且未含區塊標題文字
chart_xmls = [n for n in z.namelist() if n.startswith("xl/charts/chart")]
polluted = []
for n in chart_xmls:
    cx = z.read(n).decode("utf-8")
    if "前瞻估值" in cx or "反推目標價" in cx or "假設區" in cx:
        polluted.append(n)
assert not polluted, f"圖表被區塊標題污染: {polluted}"

print("OK — 橫向版面 / 目標價修正 / row1 淨空 / 凍結 C9 / C-F 隱藏 / 5 圖 / ignoredErrors 全過")
print("目標價(P/S) 公式:", tps)
