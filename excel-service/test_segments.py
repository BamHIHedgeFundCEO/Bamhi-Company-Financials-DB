"""煙霧測試：合成分部 payload → 生成活頁簿 → 讀回驗證分部分頁的硬規則。

鎖住這些容易做錯的不變式：
  1. 上層匯總（Apple 的「產品」含 iPhone/Mac/iPad）不能進合計，否則重複計算
  2. 合計與各項比率一律是 Excel 公式，不是算好的數值
  3. 各公司揭露科目不同 → 列是動態長出來的，不是寫死
  4. 年度欄與季度欄分段，且圖表只取其中一段（混畫會讓年度長條比季度高四倍）
  5. 季別由會計年度結束月份回推，不是看曆月

python test_segments.py
"""
from io import BytesIO

from openpyxl import load_workbook

from workbook import _period_label, _structure_generations, build_workbook

# 會計年度 9 月結（Apple 式）：年度兩期 + 連續四季
ANNUAL = ["2024-09-28#A", "2025-09-27#A"]
QUARTERS = ["2025-12-27#Q", "2026-03-28#Q", "2026-06-27#Q", "2026-09-26#Q"]
PERIODS = ANNUAL + QUARTERS


def cell(v, verified=True, parent=False):
    return {"value": v, "verified": verified, "isParent": parent}


def member(key, zh, en, vals, verified=True):
    """vals: {period: {concept: (值, 該欄是否為上層)}}；verified 三態，None = 無法校驗"""
    values = {}
    for p, byc in vals.items():
        values[p] = {c: cell(v, verified=verified, parent=par) for c, (v, par) in byc.items()}
    return {"key": key, "zh": zh, "en": en, "values": values}


SEGMENTS = {
    "company": "Test Co", "cik": "0000000001", "ticker": "TEST",
    "configVersion": "1.0", "periods": PERIODS, "warnings": [],
    "axes": [{
        "axis": "srt:ProductOrServiceAxis", "role": "product",
        "zh": "產品與服務", "en": "Product / Service",
        "concepts": ["revenue", "cogs"],
        "members": [
            # 子項：加總 = 合併總額
            member("iphone", "iPhone", "iPhone",
                   {p: {"revenue": (200, False)} for p in PERIODS}),
            member("mac", "Mac", "Mac",
                   {p: {"revenue": (100, False)} for p in PERIODS}),
            member("service", "服務", "Service",
                   {p: {"revenue": (100, False), "cogs": (25, False)} for p in PERIODS}),
            # 上層：本身含 iPhone + Mac，成本只揭露到這一層
            member("product", "產品", "Product",
                   {p: {"revenue": (300, True), "cogs": (200, False)} for p in PERIODS}),
            # ORCL 式：年報多揭露一層（此列在年度欄是上層），10-Q 只揭露這一層
            # （季度欄它就是子項）。整列抽掉會讓季度欄少算，所以留在子項區塊逐欄扣。
            member("region", "大區", "Region",
                   {p: {"revenue": (400, p in ANNUAL)} for p in PERIODS}),
            # 無法校驗（ASC 280 自訂分部利潤定義）→ 不該被標橘底
            member("unverifiable", "無法校驗項", "Unverifiable",
                   {p: {"revenue": (1, False)} for p in PERIODS}, verified=None),
            # 真的對不上 → 必須標橘底
            member("mismatch", "對不上項", "Mismatch",
                   {p: {"revenue": (2, False)} for p in PERIODS}, verified=False),
        ],
    }],
}

FIN = {
    "company": "Test Co", "ticker": "TEST", "cik": "0000000001", "mapVersion": "test",
    "periodicity": "quarterly", "currency": "USD", "periods": ["FY2025 Q1"],
    "lineItems": [], "derived": [],
}


def main() -> None:
    data = build_workbook({"financials": FIN, "segments": SEGMENTS})
    wb = load_workbook(BytesIO(data))
    assert "分部數據" in wb.sheetnames, "分部分頁沒生成"
    ws = wb["分部數據"]

    col = {}  # A 欄標籤 → 列號
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str):
            col.setdefault(v.strip(), r)

    # 1. 上層匯總必須被標示，且排在合計之後
    parent_label = next(k for k in col if k.startswith("產品（上層匯總"))
    total_rows = [r for k, r in col.items() if k == "合計"]
    assert total_rows, "沒有合計列"
    rev_total = min(total_rows)
    assert col[parent_label] > rev_total, "上層匯總必須排在合計之後"

    # 2. 合計是 SUM 公式，且範圍不含「每一欄都是上層」那一列
    import re

    def total_formula(c):
        f = ws.cell(row=rev_total, column=c).value
        assert isinstance(f, str) and f.startswith("=SUM("), f"合計不是公式：{f}"
        return f

    f = total_formula(3)
    lo, hi = (int(x) for x in re.match(r"=SUM\(C(\d+):C(\d+)\)", f).groups())
    assert not (lo <= col[parent_label] <= hi), f"上層被算進合計了：{f}"

    # 3. 合計範圍涵蓋全部子項（iPhone/Mac/服務/大區/無法校驗項/對不上項）
    assert hi - lo + 1 == 6, f"合計應涵蓋 6 個子項，實際 {hi - lo + 1}"
    vals = [ws.cell(row=r, column=3).value for r in range(lo, hi + 1)]
    assert sum(vals) == 803, f"子項加總應為 803，實際 {vals}"

    # 3a. 只在部分欄是上層的成員（大區）：年度欄要從合計扣掉、季度欄不能扣。
    #     跨欄共用固定 SUM 範圍時這件事表達不出來，年度欄就會多算一整層。
    mixed_label = next(k for k in col if k.startswith("大區（部分期間"))
    mixed_row = col[mixed_label]
    assert lo <= mixed_row <= hi, "部分期間為上層的成員應留在子項區塊"
    annual_f, quarter_f = total_formula(3), total_formula(5)
    assert annual_f.endswith(f"-C{mixed_row}"), f"年度欄沒扣掉該欄的上層：{annual_f}"
    assert "-E" not in quarter_f, f"季度欄不該扣（那一欄它是子項）：{quarter_f}"

    # 3b. verified 三態：只有 False 標橘底，None（無法校驗）不標
    def has_fill(label):
        c = ws.cell(row=col[label], column=3)
        return c.fill is not None and c.fill.fgColor.rgb not in (None, "00000000")

    assert has_fill("對不上項"), "verified=False 必須標橘底"
    assert not has_fill("無法校驗項"), "verified=None 是無法校驗，不該標橘底"
    assert not has_fill("iPhone"), "verified=True 不該標橘底"

    # 4. 比率一律公式（IFERROR 包除法），不能是算好的數值
    for label in ("營收佔比", "分部毛利率"):
        assert label in col, f"缺少 {label} 區塊"
        r = col[label] + 1
        v = ws.cell(row=r, column=3).value
        assert isinstance(v, str) and v.startswith("=IFERROR("), f"{label} 不是公式：{v}"

    # 5. 上層才有成本 → 產品毛利率要算得出來（(300-200)/300），這正是 Apple 的情境
    gm_rows = [r for r in range(col["分部毛利率"] + 1, ws.max_row + 1)
               if isinstance(ws.cell(row=r, column=1).value, str)
               and ws.cell(row=r, column=1).value.strip() == "產品"]
    assert gm_rows, "上層有揭露成本，卻沒算出毛利率"

    # 6. 營收佔比不含上層，否則佔比總和會超過 100%
    share_labels = []
    r = col["營收佔比"] + 1
    while isinstance(ws.cell(row=r, column=1).value, str) and ws.cell(row=r, column=3).value:
        share_labels.append(ws.cell(row=r, column=1).value.strip())
        r += 1
    assert "產品" not in share_labels, f"上層不該出現在營收佔比：{share_labels}"

    # 7. 欄位標籤：年度給 FY、季度回推季別（9 月結 → 6 月底是 Q3，不是 Q2）
    labels = [ws.cell(row=1, column=3 + i).value for i in range(len(PERIODS))]
    assert labels[:2] == ["FY2024", "FY2025"], labels
    assert labels[2:] == ["FY2026 Q1", "FY2026 Q2", "FY2026 Q3", "FY2026 Q4"], labels

    # 8. 圖表要畫出來，且只涵蓋季度欄（第 5 欄起共 4 欄），不能把年度混進去
    assert ws._charts, "分部分頁沒有圖表"
    seen = set()
    for ch in ws._charts:
        for s in ch.series:
            ref = s.val.numRef.f  # 例：'分部數據'!$E$5:$H$5
            rng = ref.split("!")[1].replace("$", "")
            lo, hi = rng.split(":")
            seen.add((lo[0], hi[0]))
    assert seen == {("E", "H")}, f"圖表欄範圍應只含 4 個季度欄 E:H，實際 {seen}"

    print("OK — 上層不進合計 / 合計為公式 / 比率為公式 / 上層毛利率保留 / "
          "佔比排除上層 / 季別回推正確 / 圖表只取季度欄")


def test_period_label() -> None:
    """季別要靠會計年度結束月份回推。看曆月的話 NVDA 會全錯一整年。"""
    # NVDA：1 月結 → 4 月底是「下一個會計年度」的 Q1
    assert _period_label("2026-04-26#Q", 1) == "FY2027 Q1"
    assert _period_label("2026-01-25#A", 1) == "FY2026"
    # AAPL：9 月結
    assert _period_label("2026-06-27#Q", 9) == "FY2026 Q3"
    assert _period_label("2025-12-27#Q", 9) == "FY2026 Q1"
    # 曆年公司：12 月結
    assert _period_label("2026-03-31#Q", 12) == "FY2026 Q1"
    # 推不出會計年度結束月 → 保守標法，不猜季別
    assert _period_label("2026-06-30#Q", None) == "2026-06 季"
    print("OK — 季別由會計年度結束月份回推（NVDA 1 月結 / AAPL 9 月結皆正確）")


def _axis(members):
    return {"zh": "軸", "en": "Axis", "concepts": ["revenue"], "members": members}


def _m(key, periods_with_data):
    return {"key": key, "zh": key, "en": key,
            "values": {p: {"revenue": cell(1)} for p in periods_with_data}}


def test_structure_generations() -> None:
    """
    分代只在「成員標籤整組換掉」時發生，不能被一般的分部增減觸發。

    這條線是 BW 帶出來的：它在 FY2025 10-K 把三個分部併成單一分部，新舊成員互斥，
    整張表半數格子變 n/a、合計的 SUM 還橫跨兩代。但同樣的機制若太敏感，
    只是新增一個分部就會把表切兩半 —— 兩個方向都要鎖住。
    """
    ps = PERIODS  # 2 年度 + 4 季

    # ① 整組換名（BW）：期間互斥 → 切兩代，且兩代的欄位不重疊
    gens = _structure_generations(
        _axis([_m("old_a", ps[:1]), _m("old_b", ps[:1]),
               _m("new_a", ps[1:]), _m("new_b", ps[1:])]), ps)
    assert len(gens) == 2, f"整組換名應切成兩代，實際 {len(gens)}"
    assert gens[0][1] & gens[1][1] == frozenset(), "兩代的欄位必須互斥"
    assert [sorted(m["key"] for m in g[0]) for g in gens] == [
        ["old_a", "old_b"], ["new_a", "new_b"]], "應依最早出現的欄排序"

    # ② 只是多開一個分部：舊成員的期間涵蓋新成員 → 相交 → 不切
    gens = _structure_generations(
        _axis([_m("always", ps), _m("added_late", ps[3:])]), ps)
    assert len(gens) == 1, "新增分部不該把表切成兩代"

    # ③ 裁撤一個分部：同理不切
    gens = _structure_generations(
        _axis([_m("always", ps), _m("dropped_early", ps[:2])]), ps)
    assert len(gens) == 1, "裁撤分部不該把表切成兩代"

    # ④ 只有年度揭露的軸（地區軸的常態）：未切代，但回傳的欄位只含年度欄，
    #    季度欄才認得出是「不適用」而不是「該有卻沒有」
    gens = _structure_generations(_axis([_m("us", ANNUAL), _m("cn", ANNUAL)]), ps)
    assert len(gens) == 1
    assert gens[0][1] == frozenset({0, 1}), f"只有年度揭露時欄位應只含年度欄，實際 {gens[0][1]}"

    # ⑤ 零散揭露（每個成員各自一期，互不相交）不是改組，不切
    gens = _structure_generations(
        _axis([_m(f"one_{i}", [p]) for i, p in enumerate(ps)]), ps)
    assert len(gens) == 1, "分出 4 代以上視為零散揭露，不該切"
    print("OK — 改組才分代；新增／裁撤分部與零散揭露都不分代")


def test_inapplicable_not_na() -> None:
    """
    不適用的格子要寫「—」而不是 n/a，且合計不能跨代 SUM。

    兩者混成同一個符號的話，使用者看到滿版 n/a 會以為資料抓漏了；
    合計跨代則是：只要哪一期兩代都揭露就會重複計算。
    """
    ps = PERIODS
    seg = dict(SEGMENTS, axes=[{
        "axis": "us-gaap:StatementBusinessSegmentsAxis", "role": "business",
        "zh": "營運分部", "en": "Business Segments", "concepts": ["revenue"],
        "members": [
            member("old1", "舊一", "Old 1", {p: {"revenue": (10, False)} for p in ps[:1]}),
            member("old2", "舊二", "Old 2", {p: {"revenue": (20, False)} for p in ps[:1]}),
            member("new1", "新一", "New 1", {p: {"revenue": (30, False)} for p in ps[1:]}),
        ],
    }])
    wb = load_workbook(BytesIO(build_workbook(
        {"cacheKey": "t.xlsx", "financials": FIN, "segments": seg})))
    ws = wb["分部數據"]

    titles = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
    assert any(t and "報導結構 1／2" in t for t in titles), f"沒有分代標題：{titles}"

    grid = [[ws.cell(r, c).value for c in range(3, 3 + len(ps))]
            for r in range(2, ws.max_row + 1)]
    assert not any(v == "n/a" for row in grid for v in row), "不適用的格子不該寫 n/a"
    assert any(v == "—" for row in grid for v in row), "不適用的格子應寫「—」"

    # 舊結構的合計只涵蓋舊成員；不適用的欄留白，不能 SUM 出 0
    old_total = next(row for row in grid
                     if isinstance(row[0], str) and row[0].startswith("=SUM("))
    assert old_total[0] == "=SUM(C5:C6)", f"合計範圍跨代了：{old_total[0]}"
    assert old_total[1] == "—", f"不適用的欄不該放 SUM 公式（會算出 0）：{old_total[1]}"
    print("OK — 不適用寫「—」、合計不跨代、空欄不 SUM 出假 0")


def test_api_layer() -> None:
    """
    打真正的 /generate 端點，不是直接呼叫 build_workbook。

    這條測試存在的理由：Pydantic 的 GeneratePayload 若沒宣告 segments，會把它
    靜默丟棄——不報錯、不警告，就是無聲消失。曾經因此線上少了整個分部分頁，
    而所有直接呼叫 build_workbook 的測試都是綠的。
    """
    from fastapi.testclient import TestClient

    from main import app

    res = TestClient(app).post(
        "/generate", json={"cacheKey": "t.xlsx", "financials": FIN, "segments": SEGMENTS})
    assert res.status_code == 200, f"{res.status_code} {res.text[:200]}"
    wb = load_workbook(BytesIO(res.content))
    assert "分部數據" in wb.sheetnames, "segments 在 API 層被丟掉了（Pydantic 未宣告欄位）"
    print("OK — /generate 端點確實把 segments 傳到 build_workbook")


if __name__ == "__main__":
    test_period_label()
    main()
    test_structure_generations()
    test_inapplicable_not_na()
    test_api_layer()
