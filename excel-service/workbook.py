"""
每次從零生成整本活頁簿（不用範本檔——openpyxl 重存會遺失圖表）。
分頁：說明 / 損益表 / 資產負債表 / 現金流量表 / 關鍵指標 / 估值倍數 / 分部數據 / 原始資料
（估值倍數與分部數據為條件式：payload 沒帶就不產生）。
版面：A 欄中文、B 欄英文、C 欄起季度；凍結 C2；缺值 n/a 絕不寫 0；
Q4 推算值淺橘底；關鍵指標全公式。
"""
from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from charts import build_range_chart, place_charts, set_data_start
from config_loader import chart_spec, theme, xbrl_map
from formulas import FIRST_DATA_COL, RefResolver, translate

STATEMENT_SHEETS = {"IS": "損益表", "BS": "資產負債表", "CF": "現金流量表"}
METRICS_SHEET = "關鍵指標"

DISCLAIMER = (
    "資料來源為美國 SEC EDGAR 公開資料（companyfacts XBRL API），本工具與 SEC 無任何隸屬關係。"
    "缺值以 n/a 表示——SEC 無該標籤不代表數值為零。僅供參考，不構成投資建議。"
)


def _fmt(unit: str, th: dict) -> str:
    nf = th["number_formats"]
    return {
        "USD": nf["usd"],
        "shares": nf["shares"],
        "USD/shares": nf["per_share"],
    }.get(unit, nf["usd"])


def _metric_fmt(formula: str, mid: str, th: dict) -> str:
    nf = th["number_formats"]
    if mid in ("dso", "dio", "dpo", "ccc"):
        return nf["days"]
    if mid in ("ocf_to_net_income", "net_debt_to_ebitda", "interest_coverage",
               "current_ratio", "quick_ratio", "debt_to_equity", "asset_turnover"):
        return nf["multiple"]
    if mid in ("ebitda", "net_debt", "fcf", "shareholder_return"):
        return nf["usd"]
    if mid in ("revenue_per_share", "book_value_per_share", "cash_plus_sti_per_share",
               "fcf_per_share", "ocf_per_share"):
        return nf["per_share"]  # 每股金額 0.00
    return nf["ratio"]


_THIN = Side(style="thin", color="E3E5E1")
_INK = Side(style="medium", color="15171A")
_BOX = Side(style="thin", color="8C9199")
ROW_BORDER = Border(bottom=_THIN)
HEADER_BORDER = Border(bottom=_INK)
# 全框（說明分頁表格用，四邊細灰框）
CELL_BOX = Border(left=_BOX, right=_BOX, top=_BOX, bottom=_BOX)


def _header_cell(ws, col: int, text: str, th: dict):
    c = ws.cell(row=1, column=col, value=text)
    c.font = Font(bold=True, size=th["fonts"]["header_size"], color=th["palette"]["header_font"].lstrip("#"))
    c.fill = PatternFill("solid", fgColor=th["palette"]["header_fill"].lstrip("#"))
    c.alignment = Alignment(horizontal="center")
    c.border = HEADER_BORDER


def _init_sheet(ws, periods: list[str], th: dict, tab_color: str | None = None):
    _header_cell(ws, 1, "科目", th)
    _header_cell(ws, 2, "Line Item", th)
    for i, p in enumerate(periods):
        _header_cell(ws, FIRST_DATA_COL + i, p, th)
    ws.freeze_panes = th["layout"]["freeze_panes"]  # C2
    ws.sheet_view.showGridLines = False
    if tab_color:
        ws.sheet_properties.tabColor = tab_color.lstrip("#")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 30
    for i in range(len(periods)):
        ws.column_dimensions[get_column_letter(FIRST_DATA_COL + i)].width = 14


def build_workbook(payload: dict) -> bytes:
    fin = payload["financials"]
    periods: list[str] = fin["periods"]
    n = len(periods)
    annual = fin.get("periodicity") == "annual"
    lookback = int(fin.get("lookbackCount", 0))  # 前面幾欄是 lookback（隱藏，供 YoY/TTM 公式）
    n_display = n - lookback
    data_start = FIRST_DATA_COL + lookback  # 圖表與顯示的第一欄
    set_data_start(data_start)
    pre_ipo = fin.get("preIpoBefore")  # 上市/借殼前的期界線
    th = theme()
    cmap = xbrl_map()
    spec = chart_spec()["sheets"]

    q4_fill = PatternFill("solid", fgColor=th["palette"]["q4_estimated_fill"].lstrip("#"))
    missing = th["layout"]["missing_value"]  # "n/a"

    wb = Workbook()
    wb.remove(wb.active)

    # ── 1. 說明 ─────────────────────────────────────────────
    info = wb.create_sheet("說明")
    info.sheet_view.showGridLines = False
    info.sheet_properties.tabColor = th["palette"]["header_font"].lstrip("#")
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 64
    info.column_dimensions["C"].width = 46
    info.column_dimensions["D"].width = 90
    # 標題列
    t = info.cell(row=1, column=1, value=f"{fin['ticker']}　{fin['company']}")
    t.font = Font(bold=True, size=16)
    info.cell(row=2, column=1, value="BamHI 美股財報庫　·　SEC EDGAR 官方資料").font = Font(
        size=10, color="8C9199")
    meta_rows = [
        ("公司", fin["company"]),
        ("Ticker", fin["ticker"]),
        ("CIK", fin["cik"]),
        ("期間", f"{periods[0]} ～ {periods[-1]}" if periods else "—"),
        ("頻率", "年度（外國發行人 20-F，無季報）" if annual else "季度"),
        ("幣別", fin.get("currency", "USD")),
        ("資料來源", "SEC EDGAR companyfacts XBRL API"),
        ("生成時間 (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")),
        ("對照表版本", fin["mapVersion"]),
        ("「推算」是什麼", "—（年度資料無推算）" if annual
         else "美股公司一年只申報 3 份 10-Q + 1 份 10-K：第四季沒有單獨的季報，"
              "SEC 上不存在「Q4 單季」這個數字。本檔以 全年 − Q1 − Q2 − Q3 計算出 Q4，"
              "數學上準確、非估計；橘底僅提醒「此數字不是直接抄自某份財報」。"
              "現金流量表的 Q2/Q3 亦由累計值差分還原（10-Q 只申報年初至今累計）。"),
        ("「n/a」是什麼", "該公司該期間沒有申報此科目——通常是公司本來就沒有這個項目"
         "（如未配息、未買回庫藏股、費用未拆分），不代表數值為零。可對照「原始資料」分頁查證。"),
        *([("上市前資料", f"{pre_ipo} 之前的季度已顯示為 n/a。此公司經 SPAC 借殼／IPO 上市，"
            "上市前為私有公司，股數基礎與上市後不可比（每股數值會嚴重失真），故不列出。")]
          if pre_ipo else []),
        *([("分部數據", "各分部（事業別／產品別／地區別）的營收與獲利。這些數字不在 SEC 的 "
            "companyfacts API 裡——該 API 不含維度資料——而是直接從申報的 XBRL instance 檔解析。"
            "揭露哪些科目由各公司依 ASC 280 自行決定（主要營運決策者看什麼才揭露什麼），"
            "所以每家公司的列數不同。標「上層匯總」的列不計入合計，避免重複計算"
            "（如 Apple 的「產品」本身已含 iPhone／Mac／iPad）。"
            "欄位左段為年度（FY）、右段為單季（FY＋Q，均為單季而非累計）；"
            "兩段不畫在同一張圖上，否則年度長條會比季度長條高約四倍，看起來像業績暴跌。")]
          if payload.get("segments", {}).get("axes") else []),
        ("圖表", "各報表分頁：前段為 chart_spec.json 定義的組合圖，後段為每一科目各一張圖。"
         "分部數據分頁：每個分部軸各一組（營收堆疊圖＋佔比／毛利率／營業利益率折線圖），"
         "定義在同一份設定檔的 segment_charts。"
         "要自訂圖表，修改 repo 的 config/chart_spec.json 即可，不需改程式。"),
        ("免責聲明", DISCLAIMER),
    ]
    meta_fill = PatternFill("solid", fgColor=th["palette"]["header_fill"].lstrip("#"))
    for i, (k, v) in enumerate(meta_rows, start=4):
        kc = info.cell(row=i, column=1, value=k)
        kc.font = Font(bold=True)
        kc.border = CELL_BOX
        kc.fill = meta_fill
        kc.alignment = Alignment(vertical="top")
        vc = info.cell(row=i, column=2, value=v)
        vc.border = CELL_BOX
        vc.alignment = Alignment(wrap_text=True, vertical="top")
    r = len(meta_rows) + 5
    info.cell(row=r, column=1, value="指標定義總表").font = Font(bold=True, size=12)
    info.cell(row=r + 1, column=4, value="資料來源：SEC EDGAR companyfacts XBRL。每個數字對應標籤見「原始資料」分頁。").font = Font(
        size=9, color="8C9199", italic=True)
    r += 1
    for j, h in enumerate(["指標", "英文", "定義（公式）", "判讀說明"], start=1):
        c = _header_cell_at(info, r, j, h, th)
    for m in fin["derived"]:
        r += 1
        for j, val in enumerate([m["zh"], m["en"], m["formula"], m["desc"]], start=1):
            c = info.cell(row=r, column=j, value=val)
            c.border = CELL_BOX
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if j == 1:
                c.font = Font(bold=True)

    # ── 2–4. 三大報表 ───────────────────────────────────────
    resolver = RefResolver()
    # 全活頁簿 id → (worksheet, row)：圖表可跨分頁引用（如損益表圖的毛利率折線在關鍵指標分頁）
    locations: dict[str, tuple] = {}
    chart_jobs: list[tuple] = []  # (ws, specs, anchor_row) — 指標列建立後才放圖
    raw_rows: list[tuple] = []  # 原始資料分頁

    tab_colors = {"IS": "1F3A5F", "BS": "5C7699", "CF": "0E6B5A"}
    na_font = Font(color="8C9199", size=th["fonts"]["size"])
    for stmt, sheet_name in STATEMENT_SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        _init_sheet(ws, periods, th, tab_color=tab_colors[stmt])
        auto_specs = []  # 每個有資料的科目各一張圖（在 chart_spec 的組合圖之後）
        row = 1
        for li in fin["lineItems"]:
            if li["statement"] != stmt:
                continue
            row += 1
            locations[li["id"]] = (ws, row)
            resolver.add(li["id"], sheet_name, row)
            ws.cell(row=row, column=1, value=li["zh"]).border = ROW_BORDER
            ws.cell(row=row, column=2, value=li["en"]).border = ROW_BORDER
            for i, p in enumerate(periods):
                cell = ws.cell(row=row, column=FIRST_DATA_COL + i)
                cell.border = ROW_BORDER
                v = li["values"].get(p)
                if v is None or v.get("value") is None:
                    cell.value = missing  # 絕不寫 0
                    cell.font = na_font
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.value = v["value"]  # 原始美元，不除以百萬
                    cell.number_format = _fmt(li["unit"], th)
                    if v.get("isEstimated"):
                        cell.fill = q4_fill
                    raw_rows.append((li["zh"], li["en"], p, v["value"],
                                     v.get("sourceTag"), v.get("accessionOrForm"),
                                     v.get("origFiled"), v.get("filed"), v.get("endDate"), li["unit"],
                                     "Q4推算＝全年−前三季" if v.get("isEstimated") else "財報直接申報值"))
            if any(vv.get("value") is not None for vv in li["values"].values()):
                yf = {"USD": "money", "shares": None, "USD/shares": None}.get(li["unit"], "money")
                auto_specs.append({
                    "type": "line" if li["unit"] == "USD/shares" else "bar",
                    "title": f"{li['zh']} {li['en']}",
                    "series": [li["id"]],
                    "y_format": yf,
                })
        chart_jobs.append((ws, spec.get(sheet_name, []) + auto_specs, row + 3))

    # ── 5. 關鍵指標（全公式）────────────────────────────────
    ws = wb.create_sheet(METRICS_SHEET)
    _init_sheet(ws, periods, th, tab_color=th["palette"]["accent"])
    group_fill = PatternFill("solid", fgColor=th["palette"]["header_fill"].lstrip("#"))
    row = 1
    current_group = None
    group_members: dict[str, list[str]] = {}
    for m in fin["derived"]:
        if m["group"] != current_group:
            current_group = m["group"]
            row += 1
            for j in range(1, FIRST_DATA_COL + n):
                gc = ws.cell(row=row, column=j)
                gc.fill = group_fill
                gc.border = HEADER_BORDER
            ws.cell(row=row, column=1, value=f"▍{current_group}").font = Font(
                bold=True, color=th["palette"]["accent"].lstrip("#"))
        group_members.setdefault(m["group"], []).append(m["id"])
        row += 1
        locations[m["id"]] = (ws, row)
        resolver.add(m["id"], METRICS_SHEET, row)
        name_cell = ws.cell(row=row, column=1, value=m["zh"])
        name_cell.border = ROW_BORDER
        ws.cell(row=row, column=2, value=m["en"]).border = ROW_BORDER
        # 滑鼠移入指標名稱顯示判讀說明（desc）
        name_cell.comment = Comment(m["desc"], "BamHI", height=160, width=360)
        fmt = _metric_fmt(m["formula"], m["id"], th)
        for i in range(n):
            col = FIRST_DATA_COL + i
            cell = ws.cell(row=row, column=col)
            cell.border = ROW_BORDER
            if m["id"] == "ccc":
                # 現金轉換循環：存貨不適用（軟體公司）時 DIO 以 0 計，仍算得出 DSO−DPO
                dso, dio, dpo = (resolver.cell(x, col) for x in ("dso", "dio", "dpo"))
                f = (f'IFERROR({dso},0)+IFERROR({dio},0)-IFERROR({dpo},0)'
                     if dso and dpo else None)
            else:
                f = translate(m["formula"], resolver, col, annual=annual)
            if f is None:
                cell.value = missing
                cell.font = na_font
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.value = f"={f}"
                cell.number_format = fmt
    # 關鍵指標圖表：① chart_spec 重點比較圖
    #   ② 組內比較——但只把「同一種單位」的指標放同一張圖（比率/天數各自比較），
    #      金額型指標（EBITDA、淨負債、FCF…）數量級差太多，混在一起會把比率壓成平線
    #   ③ 每個指標各自一張獨立折線
    fmt_of = {m["id"]: _metric_fmt(m["formula"], m["id"], th) for m in fin["derived"]}
    nf = th["number_formats"]
    # 每種數字格式 → 顯示名 + 圖表 Y 軸單位
    bucket = {
        nf["ratio"]: ("比率", "percent"),
        nf["days"]: ("週轉天數", "days"),
        nf["multiple"]: ("倍數", "multiple"),
        nf["usd"]: ("金額", "money"),
    }
    yfmt_of = {mid: bucket.get(fmt, ("", None))[1] for mid, fmt in fmt_of.items()}
    # 只保留 ① chart_spec 精選比較圖（2-3 條，可控）② 每個指標各自一張（自動縮放、看得清）。
    # 移除「組內全指標」比較圖：虧損股的極端比率（如利息保障 -106x）會把其他線壓成一團。
    metric_specs = list(spec.get(METRICS_SHEET, []))
    for m in fin["derived"]:
        metric_specs.append(
            {"type": "line", "title": f"{m['zh']} {m['en']}", "series": [m["id"]], "y_format": yfmt_of[m["id"]]})
    chart_jobs.append((ws, metric_specs, row + 3))

    # ── 5b. 估值倍數（模型：股價為輸入格，倍數全公式）─────────
    valuation = fin.get("valuation")
    if valuation and valuation.get("rows"):
        _build_valuation_sheet(wb, valuation, locations, periods, th,
                               data_start, n_display, chart_jobs)

    # ── 5c. 分部數據（companyfacts 無維度，資料來自 XBRL instance）────
    segments = payload.get("segments")
    if segments and segments.get("axes"):
        _build_segment_sheet(wb, segments, th)

    # 指標列已定位，統一放圖（圖表可跨分頁引用系列）；收集圖表位置做「說明」目錄
    locate = locations.get
    chart_index: list = []
    for job_ws, job_specs, anchor in chart_jobs:
        place_charts(job_ws, job_specs, locate, n_display, anchor_row=anchor, index=chart_index)

    # 「說明」分頁頂端加「圖表快速跳轉」目錄：點連結直接跳到該圖，不用一路往下滑
    from openpyxl.utils import get_column_letter as _gcl
    jump_col = 6  # F 欄，避開左側 meta/指標定義
    info.cell(row=1, column=jump_col, value="圖表快速跳轉").font = Font(bold=True, size=12)
    info.column_dimensions[_gcl(jump_col)].width = 34
    ji = 2
    cur_sheet = None
    for sheet_name, title, r in chart_index:
        if sheet_name != cur_sheet:
            cur_sheet = sheet_name
            info.cell(row=ji, column=jump_col, value=f"▍{sheet_name}").font = Font(
                bold=True, color=th["palette"]["accent"].lstrip("#"))
            ji += 1
        link = info.cell(row=ji, column=jump_col, value=f"　{title}")
        link.hyperlink = f"#'{sheet_name}'!{_gcl(FIRST_DATA_COL)}{r}"
        link.font = Font(color="0E6B5A", underline="single", size=10)
        ji += 1

    # 每個報表/指標分頁的圖表快捷：在該頁圖表區頂端（資料表下方、圖表左側 A/B 空白處）
    # 集中放「所有圖表的文字連結」，點一下直接跳到該圖——不再每張圖各放一個回頂端。
    from collections import OrderedDict
    per_sheet: "OrderedDict[str, list]" = OrderedDict()
    for sheet_name, title, r in chart_index:
        per_sheet.setdefault(sheet_name, []).append((title, r))
    for sheet_name, items in per_sheet.items():
        sh = wb[sheet_name]
        top = items[0][1]  # 第一張圖的列 = 圖表區頂端
        # 第 1 列（凍結，永遠可見）放一個「跳到圖表目錄」
        sc = FIRST_DATA_COL + n + 1
        q = sh.cell(row=1, column=sc, value="► 跳到圖表目錄")
        q.hyperlink = f"#'{sheet_name}'!A{top}"
        q.font = Font(color="0E6B5A", underline="single", bold=True, size=10)
        sh.column_dimensions[_gcl(sc)].width = 16
        # 圖表區頂端的文字連結目錄（A 欄，縱向排列所有圖）
        hdr = sh.cell(row=top, column=1, value="■ 本頁圖表目錄")
        hdr.font = Font(bold=True, color=th["palette"]["accent"].lstrip("#"))
        back = sh.cell(row=top, column=2, value="回說明")
        back.hyperlink = f"#'說明'!{_gcl(jump_col)}1"
        back.font = Font(color="0E6B5A", underline="single", size=10)
        for k, (title, r) in enumerate(items, start=1):
            c = sh.cell(row=top + k, column=1, value=f"　{title}")
            c.hyperlink = f"#'{sheet_name}'!{_gcl(FIRST_DATA_COL)}{r}"
            c.font = Font(color="0E6B5A", underline="single", size=10)

    # ── 6. 原始資料 ─────────────────────────────────────────
    ws = wb.create_sheet("原始資料")
    headers = ["科目", "Line Item", "季別", "數值", "XBRL 標籤", "表單",
               "原始申報日", "取值來源申報日", "期末日", "單位", "備註"]
    for j, h in enumerate(headers, start=1):
        c = _header_cell_at(ws, 1, j, h, th)
        if h in ("原始申報日", "取值來源申報日"):
            c.comment = Comment(
                "原始申報日＝該期首次申報；取值來源申報日＝實際取值的那份（可能是後續重編，取 filed 最新）。"
                "兩者不同代表該數字曾被重編。", "BamHI", height=120, width=320)
    for i, rr in enumerate(raw_rows, start=2):
        for j, v in enumerate(rr, start=1):
            ws.cell(row=i, column=j, value=v)
    ws.freeze_panes = "A2"
    widths = [24, 28, 12, 16, 44, 10, 12, 14, 12, 10, 8]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # 隱藏 lookback 欄：資料在、公式引用得到（YoY/TTM 從第一顯示欄即活），但使用者看不到
    if lookback > 0:
        for sname in list(STATEMENT_SHEETS.values()) + [METRICS_SHEET, "估值倍數"]:
            if sname in wb.sheetnames:
                sh = wb[sname]
                for c in range(FIRST_DATA_COL, data_start):
                    sh.column_dimensions[get_column_letter(c)].hidden = True

    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    # 估值倍數：關掉 Excel 背景錯誤檢查（P/E、P/S 等公式左上角的綠色三角）。
    # openpyxl 3.1 不序列化 ignoredErrors，故存檔後直接改寫該分頁 XML。
    if valuation and valuation.get("rows") and "估值倍數" in wb.sheetnames:
        sheet_no = wb.sheetnames.index("估值倍數") + 1
        end_col = get_column_letter(max(data_start + n_display - 1, data_start + 3))
        data = _inject_ignored_errors(data, sheet_no, f"A1:{end_col}33")
    return data


def _inject_ignored_errors(data: bytes, sheet_no: int, sqref: str) -> bytes:
    """把 <ignoredErrors> 注入指定 worksheet XML（schema 規定在 <drawing> 之前）。"""
    import zipfile

    target = f"xl/worksheets/sheet{sheet_no}.xml"
    frag = (f'<ignoredErrors><ignoredError sqref="{sqref}" '
            f'formula="1" formulaRange="1" emptyCellReferences="1" '
            f'numberStoredAsText="1" unlockedFormula="1"/></ignoredErrors>')
    zin = zipfile.ZipFile(BytesIO(data))
    out = BytesIO()
    zout = zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED)
    for it in zin.infolist():
        raw = zin.read(it.filename)
        if it.filename == target and b"<ignoredErrors" not in raw:
            xml = raw.decode("utf-8")
            pos = xml.find("<drawing ")
            if pos == -1:
                pos = xml.rfind("</worksheet>")
            raw = (xml[:pos] + frag + xml[pos:]).encode("utf-8")
        zout.writestr(it, raw)
    zin.close()
    zout.close()
    return out.getvalue()


BLUE_INPUT = Font(color="0000FF")           # 藍字＝可改假設
BLUE_FILL = PatternFill("solid", fgColor="FFFFCC")  # 黃底
GREEN_LINK = Font(color="0E6B5A")           # 綠字＝跨表連結


def _build_valuation_sheet(wb, valuation, locations, periods, th, data_start, n_display, chart_jobs):
    """
    估值分頁＝財務模型：股價是唯一硬值（藍字輸入格），市值/PE/PS/EV… 全部公式。
    版面：假設區(前瞻用) → 反推目標價 → 歷史逐季倍數(全公式，TTM 引用含 lookback 欄)。
    """
    from openpyxl.utils import get_column_letter as g

    vws = wb.create_sheet("估值倍數")
    vws.sheet_view.showGridLines = False
    vws.sheet_properties.tabColor = "C25A18"
    vws.column_dimensions["A"].width = 26
    vws.column_dimensions["B"].width = 26
    for i in range(len(periods)):
        vws.column_dimensions[g(FIRST_DATA_COL + i)].width = 13

    dcol = g(data_start)          # 第一顯示欄字母（假設輸入放這欄，永遠可見）
    last_disp = data_start + n_display - 1
    lastcol = g(last_disp)        # 最後一季（最新）

    def sheet_row(cid):
        loc = locations.get(cid)
        return (loc[0].title, loc[1]) if loc else (None, None)

    rev_s, rev_r = sheet_row("revenue")
    ni_s, ni_r = sheet_row("net_income")
    epsd_s, epsd_r = sheet_row("eps_diluted")
    so_s, so_r = sheet_row("shares_outstanding")
    eq_s, eq_r = sheet_row("equity")
    fcf_s, fcf_r = sheet_row("fcf")
    ebitda_s, ebitda_r = sheet_row("ebitda")
    nd_s, nd_r = sheet_row("net_debt")

    def ref(sheet, row, col):
        return f"'{sheet}'!{g(col)}{row}"

    def ttm(sheet, row, col):
        return f"SUM('{sheet}'!{g(col - 3)}{row}:{g(col)}{row})"

    px = valuation.get("currentPrice")
    ps_med = None
    for r in valuation["rows"]:
        if r["id"] == "ps_vs_median":
            import re as _re
            m = _re.search(r"([\d.]+)", r.get("desc", ""))
            if m:
                ps_med = float(m.group(1))

    # ── 三大區塊：橫向並排（左右各一組「標籤｜數值」），開檔即全見不折疊 ──
    # 欄位相對 data_start 動態：假設區 A/B、前瞻估值 dcol/dcol+1、反推目標價 dcol+2/dcol+3。
    # ⚠ 估值圖 category 讀第 1 列（charts._cats_ref, min_row=1）→ 區塊一律從第 2 列起，
    #    第 1 列（dcol 起）必須淨空，否則區塊標題會變成圖表 X 軸標籤。
    accent = th["palette"]["accent"].lstrip("#")
    TITLE_F = Font(bold=True, color=accent)
    LABEL_F = Font(color="15171A")
    NOTE_F = Font(size=9, color="8C9199")

    # 配色走設定層（config/theme.json），取不到才回退既有常數
    pal = th["palette"]
    def _c(key, fallback):
        return pal.get(key, fallback).lstrip("#")
    INPUT_F = Font(color=_c("input_font", "0000FF"))
    INPUT_FILL = PatternFill("solid", fgColor=_c("input_fill", "FFFFCC"))
    LINK_F = Font(color=_c("link_font", "0E6B5A"))

    c2l, c2v = data_start, data_start + 1        # 前瞻：標籤 / 數值欄
    c3l, c3v = data_start + 2, data_start + 3    # 反推：標籤 / 數值欄
    L2, L3 = g(c2v), g(c3v)

    def vtitle(col, zh):
        vws.cell(row=2, column=col, value=zh).font = TITLE_F

    def vlab(col, row, zh):
        vws.cell(row=row, column=col, value=zh).font = LABEL_F

    def vval(col, row, val, fmt, kind="formula"):
        # 空的輸入格仍要建立並上色（FY+1/FY+2 EPS 待使用者填），只有非輸入格才略過
        if val is None and kind != "input":
            return
        c = vws.cell(row=row, column=col, value=val)
        c.number_format = fmt
        if kind == "input":
            c.font = INPUT_F
            c.fill = INPUT_FILL
        elif kind == "shares":
            c.font = LINK_F

    # 加寬區塊用到的欄（同時是歷史表前幾季欄，只加寬不縮窄）
    vws.column_dimensions["A"].width = 26
    vws.column_dimensions["B"].width = 20
    vws.column_dimensions[g(c2l)].width = 19
    vws.column_dimensions[g(c2v)].width = 18
    vws.column_dimensions[g(c3l)].width = 20
    vws.column_dimensions[g(c3v)].width = 13

    # 假設區（A 標籤 / B 數值）— 輸入格藍字黃底、股數綠字跨表
    vtitle(1, "■ 假設區　Assumptions")
    vlab(1, 3, "目前股價");         vval(2, 3, round(px, 2) if px else None, '$0.00', "input")
    vlab(1, 4, "流通股數（最新）");   vval(2, 4, f"={ref(so_s, so_r, last_disp)}" if so_s else None, "#,##0", "shares")
    vlab(1, 5, "FY+1 預估 EPS");    vval(2, 5, None, '0.00', "input")
    vlab(1, 6, "FY+2 預估 EPS");    vval(2, 6, None, '0.00', "input")
    vlab(1, 7, "目標本益比");        vval(2, 7, 20, '0.0"x"', "input")
    vlab(1, 8, "目標股價營收比");     vval(2, 8, round(ps_med, 1) if ps_med else 5, '0.0"x"', "input")

    # 前瞻估值（dcol 標籤 / dcol+1 數值）
    vtitle(c2l, "■ 前瞻估值　Forward Valuation")
    vlab(c2l, 3, "市值");            vval(c2v, 3, '=IFERROR($B$3*$B$4,"n/a")', "#,##0")
    vlab(c2l, 4, "前瞻本益比 FY+1");  vval(c2v, 4, '=IFERROR($B$3/$B$5,"n/a")', '0.0"x"')
    vlab(c2l, 5, "前瞻本益比 FY+2");  vval(c2v, 5, '=IFERROR($B$3/$B$6,"n/a")', '0.0"x"')
    vlab(c2l, 6, "預估 EPS 成長率");  vval(c2v, 6, '=IFERROR($B$6/$B$5-1,"n/a")', '0.0%')
    vlab(c2l, 7, "前瞻 PEG");         vval(c2v, 7, f'=IFERROR(${L2}$4/(${L2}$6*100),"n/a")', '0.00')
    vlab(c2l, 8, "目前本益比（TTM）"); vval(c2v, 8, f'=IFERROR($B$3/{ttm(epsd_s, epsd_r, last_disp)},"n/a")' if epsd_s else None, '0.0"x"')

    # 反推目標價（dcol+2 標籤 / dcol+3 數值）
    vtitle(c3l, "■ 反推目標價　Reverse: Target Price")
    vlab(c3l, 3, "目標價（P/E × FY+1 EPS）");  vval(c3v, 3, '=IFERROR($B$7*$B$5,"n/a")', '$0.00')
    vlab(c3l, 4, "目標價（P/E × FY+2 EPS）");  vval(c3v, 4, '=IFERROR($B$7*$B$6,"n/a")', '$0.00')
    # 目標價（P/S × TTM 營收）÷ 股數：股數引用 $B$4（原本誤用帶「=」的 p3 → 產生 /=... 壞公式）
    vlab(c3l, 5, "目標價（P/S × TTM 營收）");  vval(c3v, 5, f'=IFERROR($B$8*{ttm(rev_s, rev_r, last_disp)}/$B$4,"n/a")' if rev_s else None, '$0.00')
    vlab(c3l, 6, "上漲空間（FY+1）");          vval(c3v, 6, f'=IFERROR(${L3}$3/$B$3-1,"n/a")', '0.0%')
    vlab(c3l, 7, "上漲空間（FY+2）");          vval(c3v, 7, f'=IFERROR(${L3}$4/$B$3-1,"n/a")', '0.0%')

    # 上漲空間紅綠條件式（正綠負紅）— 粗體＋底色，跟綠字跨表連結明顯區隔
    from openpyxl.formatting.rule import CellIsRule
    up_f, up_fill = _c("upside_positive", "107C41"), _c("upside_positive_fill", "C6EFCE")
    dn_f, dn_fill = _c("upside_negative", "9C0006"), _c("upside_negative_fill", "FFC7CE")
    for rr in (6, 7):
        cc = f"{L3}{rr}"
        vws[cc].font = Font(bold=True, color=up_f)
        # ⚠ 條件格式的 dxf solid fill，Excel 讀 bgColor 而非 fgColor → 兩者都寫才會上色
        vws.conditional_formatting.add(cc, CellIsRule(
            operator="greaterThan", formula=["0"],
            font=Font(bold=True, color=up_f),
            fill=PatternFill("solid", fgColor=up_fill, bgColor=up_fill)))
        vws.conditional_formatting.add(cc, CellIsRule(
            operator="lessThan", formula=["0"],
            font=Font(bold=True, color=dn_f),
            fill=PatternFill("solid", fgColor=dn_fill, bgColor=dn_fill)))

    # 說明備註
    vws.cell(row=10, column=1, value="藍字＝可修改輸入；黑字＝公式；綠字＝跨表連結").font = NOTE_F
    vws.cell(row=11, column=1, value="FY+1 / FY+2 EPS 需自行填入（SEC 不提供分析師預估）").font = NOTE_F

    # ── 歷史逐季倍數（全公式）──
    hdr = 24
    _header_cell_at(vws, hdr, 1, "歷史逐季倍數", th)
    _header_cell_at(vws, hdr, 2, "TTM／公式", th)
    for i in range(n_display):
        col = data_start + i
        _header_cell_at(vws, hdr, col, periods[(len(periods) - n_display) + i], th)
    # 凍結：只鎖 A/B 標籤欄 + 上方區塊（第 1–8 列）；捲到下方歷史表/圖表不再被高表頭擋
    vws.freeze_panes = "C9"

    price_vals = next((r["values"] for r in valuation["rows"] if r["id"] == "price"), {})
    rows_spec = [
        ("期末股價", "Price", '$0.00', None),
        ("市值", "Market Cap", '#,##0', "mc"),
        ("本益比 P/E", "P/E (TTM)", '0.0"x"', "pe"),
        ("股價營收比 P/S", "P/S (TTM)", '0.0"x"', "ps"),
        ("股價淨值比 P/B", "P/B", '0.0"x"', "pb"),
        ("股價自由現金流比 P/FCF", "P/FCF (TTM)", '0.0"x"', "pfcf"),
        ("企業價值 EV", "Enterprise Value", '#,##0', "ev"),
        ("EV／EBITDA", "EV/EBITDA (TTM)", '0.0"x"', "eve"),
        ("PS／歷史中位數", "P/S vs Median", '0.00"倍"', "psmed"),
    ]
    row0 = hdr + 1
    rrows = {}
    for k, (zh, en, fmt, key) in enumerate(rows_spec):
        r = row0 + k
        rrows[key or "price"] = r
        vws.cell(row=r, column=1, value=zh).border = ROW_BORDER
        vws.cell(row=r, column=2, value=en).border = ROW_BORDER
    # 全期（含 lookback）都填股價，讓 TTM 公式引用得到；但只有顯示欄可見
    all_price_row = rrows["price"]
    for i, p in enumerate(periods):
        col = FIRST_DATA_COL + i
        c = vws.cell(row=all_price_row, column=col)
        c.border = ROW_BORDER
        pv = price_vals.get(p)
        if pv is not None:
            c.value = round(pv, 2)
            c.font = INPUT_F
            c.fill = INPUT_FILL
            c.number_format = '$0.00'
        else:
            c.value = missing
    pr = all_price_row
    for i in range(n_display):
        col = data_start + i
        L = g(col)
        mc = f"{L}{rrows['mc']}"
        ev = f"{L}{rrows['ev']}"
        cells = {
            "mc": f"={L}{pr}*{ref(so_s, so_r, col)}" if so_s else None,
            "pe": f'=IFERROR({mc}/{ttm(ni_s, ni_r, col)},"n/a")' if ni_s else None,
            "ps": f'=IFERROR({mc}/{ttm(rev_s, rev_r, col)},"n/a")' if rev_s else None,
            "pb": f'=IFERROR({mc}/{ref(eq_s, eq_r, col)},"n/a")' if eq_s else None,
            "pfcf": f'=IFERROR({mc}/{ttm(fcf_s, fcf_r, col)},"n/a")' if fcf_s else None,
            "ev": f"={mc}+{ref(nd_s, nd_r, col)}" if nd_s else None,
            "eve": f'=IFERROR({ev}/{ttm(ebitda_s, ebitda_r, col)},"n/a")' if ebitda_s else None,
        }
        for key, f in cells.items():
            c = vws.cell(row=rrows[key], column=col)
            c.border = ROW_BORDER
            c.number_format = dict(rows_spec_map(rows_spec))[key]
            c.value = f if f else missing
    # PS／歷史中位數：引用整條 PS 顯示範圍的 MEDIAN
    ps_row = rrows["ps"]; med_row = rrows["psmed"]
    rng = f"{g(data_start)}{ps_row}:{g(last_disp)}{ps_row}"
    for i in range(n_display):
        col = data_start + i
        c = vws.cell(row=med_row, column=col)
        c.border = ROW_BORDER
        c.number_format = '0.00"倍"'
        c.value = f'=IFERROR({g(col)}{ps_row}/MEDIAN({rng}),"n/a")'

    # 估值圖：PE/PS/PB 各一張折線（引用歷史列）
    for key in ("pe", "ps", "pb", "pfcf", "eve"):
        locations[f"val_{key}"] = (vws, rrows[key])
    vspecs = [{"type": "line", "title": t, "series": [f"val_{k}"], "y_format": None}
              for k, t in [("pe", "本益比 P/E"), ("ps", "股價營收比 P/S"), ("pb", "股價淨值比 P/B"),
                           ("pfcf", "P/FCF"), ("eve", "EV/EBITDA")]]
    chart_jobs.append((vws, vspecs, row0 + len(rows_spec) + 3))


def rows_spec_map(rows_spec):
    return [((key or "price"), fmt) for zh, en, fmt, key in rows_spec]


def _header_cell_at(ws, row: int, col: int, text: str, th: dict):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, size=th["fonts"]["header_size"], color=th["palette"]["header_font"].lstrip("#"))
    c.fill = PatternFill("solid", fgColor=th["palette"]["header_fill"].lstrip("#"))
    c.border = CELL_BOX
    return c


SEGMENT_SHEET = "分部數據"


def _split_period(key: str) -> tuple[str, str]:
    """`2025-09-27#A` → ("2025-09-27", "A")。無後綴的舊格式當年度。"""
    if "#" in key:
        end, kind = key.rsplit("#", 1)
        return end, kind
    return key, "A"


def _fy_label(iso: str) -> str:
    """期末日 → FY 標籤。以期末日年份為準（NVDA 2026-01-25 → FY2026，與公司自身命名一致）。"""
    return f"FY{iso[:4]}" if iso else "—"


def _period_label(key: str, fy_end_month: int | None) -> str:
    """
    期間 key → 欄名。年度給 `FY2025`，季度給 `FY2026 Q3`。

    季別要靠會計年度結束月份回推，不能看曆月：AAPL 的 6 月底是 Q3（會計年度
    9 月結），NVDA 的 4 月底卻是 Q1（1 月結）。fy_end_month 取自同一批資料裡的
    年度期間；真的推不出來（只有季度、沒有年度）就退回 `2026-06 季`，
    寧可標得保守也不要標錯季別。
    """
    end, kind = _split_period(key)
    if not end:
        return "—"
    if kind == "A":
        return _fy_label(end)
    m = int(end[5:7])
    if not fy_end_month:
        return f"{end[:7]} 季"
    q = ((m - fy_end_month + 11) % 12) // 3 + 1
    fy = int(end[:4]) + (0 if m <= fy_end_month else 1)
    return f"FY{fy} Q{q}"


def _build_segment_sheet(wb, seg: dict, th: dict):
    """
    分部數據分頁 —— 這裡的數字 companyfacts API 給不了。

    SEC 的 companyfacts 不含維度（dimension），分部數字只存在申報的 XBRL instance
    檔裡，由 web/server/utils/segments.ts 剖析後隨 payload 帶進來。

    **列必須是動態的**：ASC 280 規定「CODM（主要營運決策者）看什麼才揭露什麼」，
    所以各家揭露的分部科目天差地遠 —— Apple 給分部營收與成本、NVDA 給營業利益
    與折舊、銀行給稅前損益。寫死列一定會錯，因此依 payload 實際有的科目長出來。

    分部毛利率／營業利益率／營收佔比一律寫 Excel 公式（IFERROR 包除法），
    不寫算好的數值 —— 與「關鍵指標」分頁同一規則，使用者改數字時會自己重算。
    """
    cmap = xbrl_map()
    zh_of = {c["id"]: c["zh"] for c in cmap["concepts"]}
    en_of = {c["id"]: c["en"] for c in cmap["concepts"]}
    nf = th["number_formats"]
    pal = th["palette"]
    missing = th["layout"]["missing_value"]
    periods: list[str] = seg["periods"]
    ncol = len(periods)

    # 年度欄在前、季度欄在後（segments.ts 已排好）。兩段分開是刻意的：
    # 把 FY2025 的長條和 FY2026 Q1 的長條畫在同一張圖上，一根年度旁邊三根季度，
    # 高度差四倍，看起來像業績暴跌 —— 所以資料分段、圖表也各畫各的。
    kinds = [_split_period(p)[1] for p in periods]
    n_annual = sum(1 for k in kinds if k == "A")
    n_quarter = ncol - n_annual
    fy_end_month = next((int(_split_period(p)[0][5:7]) for p, k in zip(periods, kinds)
                         if k == "A"), None)

    ws = wb.create_sheet(SEGMENT_SHEET)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = pal["accent"].lstrip("#")
    _header_cell(ws, 1, "分部 / 科目", th)
    _header_cell(ws, 2, "Segment / Line Item", th)
    for i, p in enumerate(periods):
        end, kind = _split_period(p)
        _header_cell(ws, FIRST_DATA_COL + i, _period_label(p, fy_end_month), th)
        # 欄名只寫 FY / FY+季，確切期末日放註解（各家會計年度結束日不同）
        ws.cell(row=1, column=FIRST_DATA_COL + i).comment = Comment(
            f"期末日 {end}\n{'年度（12 個月）' if kind == 'A' else '單季（3 個月，非累計）'}",
            "BamHI", height=70, width=200)
    ws.freeze_panes = th["layout"]["freeze_panes"]
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 32
    for i in range(ncol):
        ws.column_dimensions[get_column_letter(FIRST_DATA_COL + i)].width = 16

    warn_fill = PatternFill("solid", fgColor=pal["q4_estimated_fill"].lstrip("#"))
    axis_font = Font(bold=True, size=12, color=pal["header_font"].lstrip("#"))
    axis_fill = PatternFill("solid", fgColor=pal["header_fill"].lstrip("#"))
    col_of = lambda i: get_column_letter(FIRST_DATA_COL + i)

    # 圖表在所有資料區之後才畫，這裡先累積 (軸中文名, spec, [(系列名, 列號)])
    pending_charts: list[tuple[str, dict, list[tuple[str, int]]]] = []
    seg_specs = chart_spec().get("segment_charts", [])

    row = 2
    for block in seg["axes"]:
        c = ws.cell(row=row, column=1, value=f"{block['zh']}")
        c.font = axis_font
        ws.cell(row=row, column=2, value=block["en"]).font = Font(
            bold=True, size=10, color=pal["header_font"].lstrip("#"))
        for j in range(1, FIRST_DATA_COL + ncol):
            ws.cell(row=row, column=j).fill = axis_fill
        row += 2

        member_rows: dict[str, dict[str, int]] = {}
        total_rows: dict[str, int] = {}
        # 科目 → 成員 → 「在哪幾欄是上層匯總」。同一個成員可能只在部分欄是上層：
        # ORCL 年報揭露 3 大區與 5 個國家兩層（各自都等於營收總額），10-Q 只有 3 大區，
        # 於是 3 大區在年度欄是上層、在季度欄是子項。合計、佔比、圖表都得逐欄判斷。
        parent_cols: dict[str, dict[str, set]] = {}
        # 圖表資料來源：只收子項（上層匯總畫進堆疊圖會讓總高度變兩倍）
        child_rows: dict[str, list[tuple[str, int, frozenset]]] = {}
        derived_rows: dict[str, list[tuple[str, int]]] = {}
        unverified = False

        for cid in block["concepts"]:
            members = [m for m in block["members"]
                       if any(cid in m["values"].get(p, {}) for p in periods)]
            if not members:
                continue
            ws.cell(row=row, column=1, value=zh_of.get(cid, cid)).font = Font(bold=True)
            ws.cell(row=row, column=2, value=en_of.get(cid, cid)).font = Font(
                bold=True, size=10, color="8C9199")
            row += 1

            def parent_at(m) -> set:
                """這個成員在哪幾欄是上層匯總（欄序號，0 起算）"""
                return {i for i, p in enumerate(periods)
                        if (m["values"].get(p, {}).get(cid) or {}).get("isParent")}

            def present_at(m) -> set:
                return {i for i, p in enumerate(periods) if m["values"].get(p, {}).get(cid)}

            pcols = {m["key"]: parent_at(m) for m in members}
            parent_cols[cid] = pcols

            # 「每一欄都是上層」的成員（如 Apple 的「產品」含 iPhone/Mac/iPad）排到
            # 子項之後並標示不計入合計。只在部分欄是上層的成員留在子項區塊，改由
            # 合計公式逐欄把它扣掉 —— 它在其他欄是真的子項，抽掉整列會讓那些欄少算。
            def always(m) -> bool:
                seen = present_at(m)
                return bool(seen) and pcols[m["key"]] >= seen

            children = [m for m in members if not always(m)]
            parents = [m for m in members if always(m)]

            def write_rows(items, suffix=""):
                nonlocal row, unverified
                for m in items:
                    # 只在部分欄是上層的成員要標出來，否則使用者會發現合計對不上眼前的加法
                    tag = suffix or ("（部分期間為上層匯總）" if pcols[m["key"]] else "")
                    ws.cell(row=row, column=1, value=f"　{m['zh']}{tag}")
                    ws.cell(row=row, column=2, value=m["en"]).font = Font(size=10, color="8C9199")
                    for i, p in enumerate(periods):
                        cell = m["values"].get(p, {}).get(cid)
                        v = ws.cell(row=row, column=FIRST_DATA_COL + i,
                                    value=cell["value"] if cell else missing)
                        v.number_format = nf["usd"]
                        # verified 是三態：None = 無法校驗，不是校驗沒過。
                        # ASC 280 允許公司自訂分部利潤定義、也允許只揭露部分分部的
                        # 費用，那種科目永遠對不上合併總額（ORCL 的分部營業利益、
                        # PFE 只有一個分部揭露營業成本）。那不是數字有問題，不標色，
                        # 否則整片橘色會讓真正對不上的欄位被淹掉。
                        if cell and cell["verified"] is False:
                            v.fill = warn_fill
                            unverified = True
                    member_rows.setdefault(cid, {})[m["key"]] = row
                    row += 1

            first = row
            write_rows(children)
            last = row - 1
            child_rows[cid] = [(m["zh"], member_rows[cid][m["key"]],
                                frozenset(pcols[m["key"]])) for m in children]

            if children:
                # 哪幾列在哪幾欄要被扣掉（部分欄才是上層的成員）
                excl_by_col: dict[int, list[int]] = {}
                for m in children:
                    for i in pcols[m["key"]]:
                        excl_by_col.setdefault(i, []).append(member_rows[cid][m["key"]])

                ws.cell(row=row, column=1, value="　合計").font = Font(bold=True)
                ws.cell(row=row, column=2, value="Total").font = Font(size=10, color="8C9199")
                for i in range(ncol):
                    L = col_of(i)
                    # 整段 SUM 再逐欄扣掉該欄的上層，比列舉子項短得多，也看得出範圍
                    f = f"=SUM({L}{first}:{L}{last})" + "".join(
                        f"-{L}{r}" for r in sorted(excl_by_col.get(i, [])))
                    t = ws.cell(row=row, column=FIRST_DATA_COL + i, value=f)
                    t.number_format = nf["usd"]
                    t.font = Font(bold=True)
                total_rows[cid] = row
                row += 1

            write_rows(parents, suffix="（上層匯總，不計入合計）")
            row += 1

        # ── 衍生指標：全部寫公式，隨上方數字連動 ──
        rev, tot_rev = member_rows.get("revenue", {}), total_rows.get("revenue")
        cogs, gp, oi = (member_rows.get("cogs", {}), member_rows.get("gross_profit", {}),
                        member_rows.get("operating_income", {}))

        def derived(title_zh: str, title_en: str, formula):
            """
            formula(member_key, col_letter, col_index) → Excel 公式字串。

            回 None 代表**該成員在該欄**沒有這個指標；整列每欄都是 None 才略過整列。
            逐欄判斷是必要的：只在部分欄是上層匯總的成員，那幾欄不能算佔比
            （會和它自己的子項重複），其他欄卻要算。
            """
            nonlocal row
            cols = list(enumerate(col_of(i) for i in range(ncol)))
            live = [k for k in rev if any(formula(k, L, i) for i, L in cols)]
            if not live:
                return
            ws.cell(row=row, column=1, value=title_zh).font = Font(bold=True)
            ws.cell(row=row, column=2, value=title_en).font = Font(size=10, color="8C9199")
            row += 1
            collected: list[tuple[str, int]] = []
            for k in live:
                label = next((m for m in block["members"] if m["key"] == k), None)
                name = label["zh"] if label else k
                ws.cell(row=row, column=1, value=f"　{name}")
                for i, L in cols:
                    c2 = ws.cell(row=row, column=FIRST_DATA_COL + i,
                                 value=formula(k, L, i) or missing)
                    c2.number_format = nf["ratio"]
                collected.append((name, row))
                row += 1
            derived_rows[title_zh] = collected
            row += 1

        if tot_rev:
            # 上層匯總不列入佔比 —— 否則子項各自佔比再加上父項，總和會超過 100%
            rev_parent_cols = parent_cols.get("revenue", {})
            derived("營收佔比", "Revenue Share",
                    lambda k, L, i: (None if i in rev_parent_cols.get(k, ())
                                     else f'=IFERROR({L}{rev[k]}/{L}{tot_rev},"{missing}")'))
        derived("分部毛利率", "Segment Gross Margin",
                lambda k, L, i: (f'=IFERROR({L}{gp[k]}/{L}{rev[k]},"{missing}")' if k in gp
                                 else f'=IFERROR(({L}{rev[k]}-{L}{cogs[k]})/{L}{rev[k]},"{missing}")'
                                 if k in cogs else None))
        derived("分部營業利益率", "Segment Operating Margin",
                lambda k, L, i: (f'=IFERROR({L}{oi[k]}/{L}{rev[k]},"{missing}")' if k in oi else None))

        # 這個軸要畫哪些圖：來源在 config/chart_spec.json 的 segment_charts，
        # 這裡只負責把「動態長出來的列」對上去
        for spec in seg_specs:
            if spec.get("source") == "concept":
                rows_for = child_rows.get(spec.get("concept", ""), [])
            else:
                rows_for = [(n, r, frozenset())
                            for n, r in derived_rows.get(spec.get("block", ""), [])]
            if rows_for:
                pending_charts.append((block["zh"], spec, rows_for))

        if unverified:
            note = ws.cell(
                row=row, column=1,
                value="橘底：該期分部加總與合併總額對不上，數字照實呈現、不調整。"
                      "（分部利潤／費用若各期都對不上，那是 ASC 280 允許公司自訂分部利潤定義、"
                      "或只揭露部分分部的費用所致，屬無法校驗，不標色。）")
            note.font = Font(size=9, color="8C9199")
            row += 1
        row += 1

    src = ws.cell(row=row + 1, column=1,
                  value="分部數據來自 SEC 申報的 XBRL instance 檔（companyfacts API 不含維度資料）。"
                        "各公司揭露的分部科目依 ASC 280 由該公司主要營運決策者所檢視的內容決定，因此列數各家不同。")
    src.font = Font(size=9, color="8C9199")
    row += 3

    # ── 圖表 ──────────────────────────────────────────────────────────────
    # 欄範圍：季度夠多就畫季度（趨勢才看得出來），否則退回年度。
    # 兩段永遠不混在同一張圖裡，理由見上方欄位排序的說明。
    if n_quarter >= 4:
        chart_first, chart_n, gran = FIRST_DATA_COL + n_annual, n_quarter, "季"
    elif n_annual >= 2:
        chart_first, chart_n, gran = FIRST_DATA_COL, n_annual, "年度"
    else:
        chart_first = chart_n = 0
        gran = ""

    if chart_n:
        hint = ws.cell(row=row, column=1,
                       value=f"以下圖表為{gran}資料（共 {chart_n} 期）；"
                             "堆疊圖不含上層匯總成員，否則總高度會重複計算。")
        hint.font = Font(size=9, color="8C9199")
        row += 2
        step = int(th["chart"]["height_rows"] * 0.7 / 0.53) + 5
        chart_cols = set(range(chart_first - FIRST_DATA_COL,
                               chart_first - FIRST_DATA_COL + chart_n))
        for axis_zh, spec, rows_for in pending_charts:
            # 圖只畫其中一段欄位，成員是不是上層要看**那一段**：ORCL 的 3 大區在年度欄
            # 是上層、在季度欄是子項，畫季度圖時它就該進堆疊圖
            series = [(n, r) for n, r, pc in rows_for if not (pc & chart_cols)]
            if not series:
                continue
            titled = dict(spec, title=f"{axis_zh}：{spec['title']}")
            chart = build_range_chart(ws, titled, series, chart_first, chart_n)
            if chart is None:
                continue
            ws.add_chart(chart, f"{get_column_letter(FIRST_DATA_COL)}{row}")
            row += step
